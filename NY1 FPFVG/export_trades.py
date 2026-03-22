#!/usr/bin/env python3
"""
export_trades.py — NY1 F.P.FVG Trade Export
Writes ny1_trades.xlsx with one tab per risk profile.
"""

import json
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

HERE      = Path(__file__).parent
JSON_PATH = HERE / 'ny1_results.json'
OUT_PATH  = HERE / 'ny1_trades.xlsx'

# ── Profile definitions ───────────────────────────────────────────────────────
PROFILES = [
    ('cashflow',          'Cashflow'),
    ('cashflow_extended', 'Cashflow+Extended'),
    ('pct_035_035',       '0.35% SL · 0.35% TP'),
    ('pct_035_025',       '0.35% SL · 0.25% TP'),
    ('pct_050_050',       '0.50% SL · 0.50% TP'),
    ('pct_100_100',       '1.00% SL · 1.00% TP'),
    ('pct_100_150',       '1.00% SL · 1.50% TP'),
    ('pct_100_200',       '1.00% SL · 2.00% TP'),
]

IS_PCT = {'pct_035_035','pct_035_025','pct_050_050','pct_100_100','pct_100_150','pct_100_200'}

RISK_PER_TRADE = 225

# ── Column definitions ────────────────────────────────────────────────────────
# (header, trade_key_or_fn, number_format, width)
BASE_COLS = [
    ('Date',        'date',        '@',      12),
    ('Time',        'time',        '@',       7),
    ('Day',         'dow_name',    '@',       6),
    ('Direction',   'direction',   '@',       9),
    ('Entry',       'entry',       '#,##0.00', 10),
    ('SL',          'stop',        '#,##0.00', 10),
    ('TP1',         'tp1',         '#,##0.00', 10),
    ('Risk Pts',    'risk_pts',    '0.00',     8),
    ('MAE %',       'mae_pct',     '0.0000"%"', 8),
    ('MFE1 %',      'mfe1_pct',    '0.0000"%"', 8),
    ('Outcome',     'outcome_main','@',        9),
    ('R',           'combined_r',  '0.00',     7),
    ('P&L $',       None,          '#,##0.00', 9),   # computed
]

EXT_COLS = [
    ('Runner Exit', 'runner_exit_price', '#,##0.00', 10),
    ('Runner',      'runner_outcome',    '@',         9),
    ('MFE2 %',      'mfe2_pct',          '0.0000"%"', 8),
]

PCT_ONLY_COLS = [
    ('Pct SL Price','pct_stop', '#,##0.00', 12),
    ('Pct TP Price','pct_tp',   '#,##0.00', 12),
]

# ── Styles ────────────────────────────────────────────────────────────────────
def hex_fill(hex_str):
    return PatternFill('solid', fgColor=hex_str)

HDR_FILL   = hex_fill('1E293B')
ALT_FILL   = hex_fill('0F172A')
BASE_FILL  = hex_fill('0A0E17')
WIN_FILL   = hex_fill('052E16')
LOSS_FILL  = hex_fill('2D0000')
OPEN_FILL  = hex_fill('1C1C2E')

HDR_FONT   = Font(name='Calibri', bold=True, color='94A3B8', size=9)
BASE_FONT  = Font(name='Calibri', color='CBD5E1', size=9)
WIN_FONT   = Font(name='Calibri', color='4ADE80', size=9)
LOSS_FONT  = Font(name='Calibri', color='F87171', size=9)
OPEN_FONT  = Font(name='Calibri', color='94A3B8', size=9)
TITLE_FONT = Font(name='Calibri', bold=True, color='E2E8F0', size=11)
STAT_FONT  = Font(name='Calibri', color='94A3B8', size=9)
STAT_VAL_FONT = Font(name='Calibri', bold=True, color='E2E8F0', size=9)
CENTER     = Alignment(horizontal='center', vertical='center')
LEFT       = Alignment(horizontal='left',   vertical='center')

thin = Side(style='thin', color='1F2937')
BORDER = Border(bottom=Side(style='thin', color='1F2937'))

SUMMARY_FILL = hex_fill('162032')


def outcome_style(outcome):
    if outcome == 'WIN':
        return WIN_FILL, WIN_FONT
    if outcome == 'LOSS':
        return LOSS_FILL, LOSS_FONT
    return OPEN_FILL, OPEN_FONT


# ── Build sheet ───────────────────────────────────────────────────────────────
def build_sheet(ws, key, label, trades, risk_stats):
    is_pct = key in IS_PCT
    is_ext = key == 'cashflow_extended'

    # Columns for this sheet
    cols = list(BASE_COLS)
    if is_ext:
        cols += EXT_COLS
    if is_pct:
        cols += PCT_ONLY_COLS

    rs = risk_stats or {}

    # ── Summary block (rows 1-3) ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    title_cell = ws.cell(row=1, column=1, value=f'NY1 F.P.FVG  ·  {label}')
    title_cell.font  = TITLE_FONT
    title_cell.fill  = hex_fill('0D1420')
    title_cell.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(6, len(cols)))

    stats = [
        ('Trades', rs.get('trades', len([t for t in trades if t.get('outcome_main') in ('WIN','LOSS')]))),
        ('Wins',   rs.get('wins', 0)),
        ('Losses', rs.get('losses', 0)),
        ('Win Rate', f"{rs.get('wins',0)/max(rs.get('trades',1),1)*100:.1f}%"),
        ('SL %',   f"{rs.get('sl_pct','—')}" if rs.get('sl_pct') is None else f"{rs['sl_pct']:.2f}%"),
        ('TP %',   f"{rs.get('tp_pct','—')}" if rs.get('tp_pct') is None else f"{rs['tp_pct']:.2f}%"),
        ('Avg Win $',  f"+${rs['avg_win_usd']:.0f}"  if rs.get('avg_win_usd')  is not None else '—'),
        ('Avg Loss $', f"-${abs(rs['avg_loss_usd']):.0f}" if rs.get('avg_loss_usd') is not None else '—'),
        ('Low Eq $',   f"${rs['min_equity_usd']:,.0f}" if rs.get('min_equity_usd') is not None else '—'),
        ('Blown',  'YES' if rs.get('blown') else 'NO'),
        ('Max W Run', rs.get('max_consec_wins', '—')),
        ('Max L Run', rs.get('max_consec_losses', '—')),
    ]

    col_idx = 1
    for lbl, val in stats:
        lc = ws.cell(row=2, column=col_idx, value=lbl)
        lc.font = STAT_FONT; lc.fill = SUMMARY_FILL; lc.alignment = CENTER
        vc = ws.cell(row=3, column=col_idx, value=val)
        vc.font = STAT_VAL_FONT; vc.fill = SUMMARY_FILL; vc.alignment = CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, len(str(lbl)) + 2)
        col_idx += 1

    # ── Blank separator row ───────────────────────────────────────────────────
    ws.row_dimensions[4].height = 6

    # ── Header row ────────────────────────────────────────────────────────────
    HDR_ROW = 5
    ws.row_dimensions[HDR_ROW].height = 18

    for ci, (hdr, _, _, width) in enumerate(cols, 1):
        cell = ws.cell(row=HDR_ROW, column=ci, value=hdr)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = width

    # Freeze pane below header
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, trade in enumerate(trades):
        row = HDR_ROW + 1 + ri
        ws.row_dimensions[row].height = 15
        outcome = trade.get('outcome_main', '')
        fill, font = outcome_style(outcome)

        for ci, (hdr, key_or_fn, num_fmt, _) in enumerate(cols, 1):
            if hdr == 'P&L $':
                r_val = trade.get('combined_r')
                val = round(r_val * RISK_PER_TRADE, 2) if r_val is not None else None
            elif key_or_fn is None:
                val = None
            else:
                val = trade.get(key_or_fn)

            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill      = fill
            cell.alignment = CENTER

            # Font colour: P&L gets green/red; direction gets colour; rest = standard
            if hdr == 'P&L $' and val is not None:
                cell.font = WIN_FONT if val > 0 else (LOSS_FONT if val < 0 else OPEN_FONT)
            elif hdr == 'Direction':
                cell.font = WIN_FONT if val == 'LONG' else LOSS_FONT
            elif hdr == 'Outcome':
                cell.font = font
            else:
                cell.font = font

            if num_fmt != '@':
                cell.number_format = num_fmt

    # Auto-filter on header row
    ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(len(cols))}{HDR_ROW}"

    # Tab colour
    ws.sheet_properties.tabColor = '10B981' if not rs.get('blown') else 'F87171'


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print(f'Loading {JSON_PATH} ...', end=' ', flush=True)
    data = json.loads(JSON_PATH.read_text())
    print('done')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for key, label in PROFILES:
        if key not in data:
            print(f'  skip {key} (not in JSON)')
            continue

        profile_data = data[key]
        trades       = profile_data.get('trades', [])
        risk_stats   = profile_data.get('risk_stats', {})

        # Only show resolved trades (WIN/LOSS) by default; include OPEN/NO_FILL too
        resolved = [t for t in trades if t.get('outcome_main') in ('WIN','LOSS')]

        print(f'  {label:<28}  {len(resolved):>5} resolved trades')

        # Safe sheet name (max 31 chars, no special chars)
        sheet_name = label[:31]
        ws = wb.create_sheet(title=sheet_name)
        build_sheet(ws, key, label, resolved, risk_stats)

    wb.save(OUT_PATH)
    print(f'\n  Saved → {OUT_PATH}')


if __name__ == '__main__':
    main()
