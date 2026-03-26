#!/usr/bin/env python3
"""
Daily Classification Engine — Standalone
==========================================
Reads M1 OHLC CSV and classifies every trading day as:
  DWP | DNP | R1 | R2 | Unclassified

Output: Excel workbook with:
  - Summary sheet  (counts + % per class)
  - DWP sheet
  - DNP sheet
  - R1 sheet
  - R2 sheet
  - Unclassified sheet

Each class sheet columns: Date | Day of Week | Classification | Reason
"""

import os, sys, glob, logging
import pandas as pd
from datetime import datetime, time as dtime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(message)s')
log = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
TICK                   = 0.25
HALF_DAY_BAR_THRESHOLD = 330
RTH_START              = dtime(9, 30)
RTH_END                = dtime(16, 0)

# ── Color palette (matches source script) ─────────────────────────────────────
DARK_BG    = '0D0F14'
GOLD       = 'F5C842'
TEAL       = '3DD9B3'
RED_CLR    = 'F5504A'
BLUE_CLR   = '5B9CF6'
WHITE      = 'E8EAF0'
MUTED      = '7A82A0'
CARD_BG    = '13161E'
RAISED_BG  = '1A1E28'
BORDER_CLR = '252A38'
ORANGE     = 'F5A623'
GREEN_CLR  = '2ECC71'

CLASS_COLORS = {
    'DWP':          TEAL,
    'DNP':          GOLD,
    'R1':           BLUE_CLR,
    'R2':           ORANGE,
    'Unclassified': MUTED,
}

CLASS_LABELS = {
    'DWP':          '📈  DWP — Directional With Pullback',
    'DNP':          '🚀  DNP — Directional No Pullback',
    'R1':           '🔄  R1  — Range 1',
    'R2':           '↩   R2  — Range 2',
    'Unclassified': '❓  Unclassified',
}

DAYS_OF_WEEK = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday',
                'Saturday', 'Sunday']

# ── Style helpers ──────────────────────────────────────────────────────────────
def fill(c):  return PatternFill('solid', start_color=c, fgColor=c)
def center(): return Alignment(horizontal='center', vertical='center')
def left():   return Alignment(horizontal='left',   vertical='center')

def tborder(c=BORDER_CLR):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def wcol(ws, col, w):
    ws.column_dimensions[get_column_letter(col)].width = w

def wc(ws, r, c, v, font=None, fill_=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font          = font
    if fill_:  cell.fill          = fill_
    if align:  cell.alignment     = align
    if border: cell.border        = border
    if fmt:    cell.number_format = fmt
    return cell

def hf(sz=11, bold=True,  color=WHITE): return Font(name='Arial', size=sz, bold=bold,  color=color)
def cf(sz=10, bold=False, color=WHITE): return Font(name='Arial', size=sz, bold=bold,  color=color)


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
def find_csv_files():
    patterns = ['*.csv', '*.txt', '*.tsv']
    found = []
    for p in patterns:
        found.extend(glob.glob(p))
        found.extend(glob.glob(os.path.join(os.path.dirname(__file__), p)))
    return list(dict.fromkeys(found))


def load_ohlc(filepath):
    with open(filepath, 'r') as f:
        sample = f.read(4096)

    sep = '\t' if sample.count('\t') > sample.count(',') else ','
    df  = pd.read_csv(filepath, sep=sep, header=0, low_memory=False)
    df.columns = [c.strip().lower() for c in df.columns]

    col_map = {}
    for col in df.columns:
        c = col.strip().lower()
        if   c in ('o', 'open'):          col_map[col] = 'open'
        elif c in ('h', 'high'):          col_map[col] = 'high'
        elif c in ('l', 'low'):           col_map[col] = 'low'
        elif c in ('c', 'close', 'last'): col_map[col] = 'close'
    df = df.rename(columns=col_map)

    raw      = [c for c in df.columns if c not in ('open', 'high', 'low', 'close')]
    dt_col   = next((c for c in raw if c in ('datetime', 'timestamp', 'dt')), None)
    date_col = next((c for c in raw if c == 'date' or ('date' in c and 'time' not in c)), None)
    time_col = next((c for c in raw if c == 'time' or ('time' in c and 'date' not in c)), None)

    for col in [date_col, time_col, dt_col]:
        if col is not None:
            df = df[df[col].notna()]
            df = df[df[col].astype(str).str.strip().str.lower() != 'nan']
    df = df.reset_index(drop=True)

    if dt_col is not None:
        df['datetime'] = pd.to_datetime(df[dt_col].astype(str))
    elif date_col is not None and time_col is not None:
        df['datetime'] = pd.to_datetime(
            df[date_col].astype(str).str.strip() + ' ' +
            df[time_col].astype(str).str.strip())
    elif date_col is not None:
        df['datetime'] = pd.to_datetime(df[date_col].astype(str))
    else:
        raise ValueError(f"Cannot find datetime columns. Found: {list(df.columns)}")

    for r in ['open', 'high', 'low', 'close']:
        if r not in df.columns:
            raise ValueError(f"Missing column '{r}'. Found: {list(df.columns)}")

    df = df[['datetime', 'open', 'high', 'low', 'close']].copy()
    df[['open', 'high', 'low', 'close']] = df[['open', 'high', 'low', 'close']].apply(
        pd.to_numeric, errors='coerce')
    df.dropna(inplace=True)
    df.sort_values('datetime', inplace=True)
    df.reset_index(drop=True, inplace=True)

    log.info(f"Loaded {len(df):,} bars | "
             f"{df['datetime'].min().date()} -> {df['datetime'].max().date()}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# CLASSIFICATION ENGINE  (exact logic from source script)
# ══════════════════════════════════════════════════════════════════════════════
def get_rth_bars(grp):
    return grp[
        (grp['datetime'].dt.time >= RTH_START) &
        (grp['datetime'].dt.time <  RTH_END)
    ].reset_index(drop=True)


def get_930_candle(grp):
    bar = grp[grp['datetime'].dt.time == dtime(9, 30)]
    if bar.empty:
        return None
    row = bar.iloc[0]
    return {'open': row['open'], 'high': row['high'],
            'low':  row['low'],  'close': row['close']}


def get_hourly_extremes(rth_bars):
    hourly = {}
    for h in range(9, 16):
        subset = rth_bars[rth_bars['datetime'].dt.hour == h]
        if subset.empty:
            continue
        hourly[h] = {'high': subset['high'].max(), 'low': subset['low'].min()}
    return hourly


def hour_touches_930(h_high, h_low, ref_high, ref_low):
    return h_low <= ref_high and h_high >= ref_low


def classify_day(grp):
    rth_bars = get_rth_bars(grp)
    if len(rth_bars) < HALF_DAY_BAR_THRESHOLD:
        return 'Unclassified', f'HALF DAY ({len(rth_bars)} RTH bars)'

    c930 = get_930_candle(grp)
    if c930 is None:
        return 'Unclassified', 'NO 9:30 BAR'

    ref_high = c930['high']
    ref_low  = c930['low']
    if ref_high == ref_low:
        ref_high += TICK
        ref_low  -= TICK

    hourly = get_hourly_extremes(rth_bars)
    if len(hourly) < 4:
        return 'Unclassified', f'INSUFFICIENT RTH HOURS ({len(hourly)})'

    hourly_close = {}
    for h in range(10, 15):
        subset = rth_bars[rth_bars['datetime'].dt.hour == h]
        if not subset.empty:
            hourly_close[h] = subset.iloc[-1]['close']

    # R1: 4+ hours (10-14) wick into 9:30
    r1_eligible = [h for h in range(10, 15) if h in hourly]
    touch_count = sum(
        1 for h in r1_eligible
        if hour_touches_930(hourly[h]['high'], hourly[h]['low'], ref_high, ref_low)
    )
    if touch_count >= 4:
        return 'R1', f'{touch_count} hours (10-14) touch 9:30 range'

    # Find first hour that closes cleanly away from 9:30
    gap_hour = None
    gap_dir  = None
    for h in sorted(hourly_close.keys()):
        if h not in hourly:
            continue
        c  = hourly_close[h]
        hh = hourly[h]['high']
        hl = hourly[h]['low']
        if c > ref_high and hl > ref_high:
            gap_hour = h; gap_dir = 'UP'; break
        elif c < ref_low and hh < ref_low:
            gap_hour = h; gap_dir = 'DOWN'; break

    if gap_hour is None:
        return 'Unclassified', f'NO HOUR CLOSED CLEANLY AWAY FROM 9:30, touch_count={touch_count}'

    # Check if any hour after gap_hour reverts to 9:30 before 15:00
    post_gap_hours = sorted([h for h in range(gap_hour + 1, 15) if h in hourly])
    revert_hour = None
    for h in post_gap_hours:
        if hour_touches_930(hourly[h]['high'], hourly[h]['low'], ref_high, ref_low):
            revert_hour = h; break

    # R2: reverts to 9:30
    if revert_hour is not None:
        return 'R2', f'Gap hour {gap_hour} ({gap_dir}), hour {revert_hour} reverts to 9:30'

    # DWP vs DNP
    all_session_hours = sorted([h for h in range(10, 15) if h in hourly])
    dwp_pullback   = False
    dwp_sweep_hour = None
    for i in range(1, len(all_session_hours)):
        prev_h = all_session_hours[i - 1]
        curr_h = all_session_hours[i]
        if gap_dir == 'UP':
            if hourly[curr_h]['low'] < hourly[prev_h]['low']:
                dwp_pullback = True; dwp_sweep_hour = curr_h; break
        else:
            if hourly[curr_h]['high'] > hourly[prev_h]['high']:
                dwp_pullback = True; dwp_sweep_hour = curr_h; break

    if dwp_pullback:
        return 'DWP', f'Gap {gap_dir} from hour {gap_hour}, hour {dwp_sweep_hour} swept prior hour extreme'
    else:
        return 'DNP', f'Gap {gap_dir} from hour {gap_hour}, no opposing extreme swept'


# ══════════════════════════════════════════════════════════════════════════════
# RUN CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════════
def run_classification(df_m1, date_range=None):
    df_m1 = df_m1.copy()
    df_m1['date'] = df_m1['datetime'].dt.date

    if date_range:
        s, e = date_range
        df_m1 = df_m1[(df_m1['date'] >= s) & (df_m1['date'] <= e)]

    results = []
    for day, grp in df_m1.groupby('date'):
        grp = grp.reset_index(drop=True)
        day_class, class_reason = classify_day(grp)
        dow = DAYS_OF_WEEK[datetime.strptime(str(day), '%Y-%m-%d').weekday()]
        results.append({
            'date':               str(day),
            'day_of_week':        dow,
            'day_classification': day_class,
            'class_reason':       class_reason,
        })

    log.info(f"Days processed: {len(results)}")
    for cls in ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']:
        cnt = sum(1 for r in results if r['day_classification'] == cls)
        pct = round(cnt / len(results) * 100, 1) if results else 0
        log.info(f"    {cls:>14}: {cnt:>5}  ({pct}%)")
    return results


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_summary_sheet(wb, results):
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_properties.tabColor = GOLD
    ws.sheet_view.showGridLines = False

    total = len(results)

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:E1')
    wc(ws, 1, 1, 'DAILY CLASSIFICATION SUMMARY',
       font=hf(14, True, DARK_BG), fill_=fill(GOLD),
       align=center(), border=tborder(DARK_BG))
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:E2')
    date_range_str = ''
    if results:
        date_range_str = f"{results[0]['date']}  →  {results[-1]['date']}   |   {total} trading days"
    wc(ws, 2, 1, date_range_str,
       font=cf(10, False, MUTED), fill_=fill(RAISED_BG),
       align=center(), border=tborder())
    ws.row_dimensions[2].height = 18

    # spacer
    ws.row_dimensions[3].height = 8

    # ── Table header ──────────────────────────────────────────────────────────
    headers = ['Classification', 'Days', 'Pct of Total', '% Bar', 'Tab Color']
    widths  = [34, 10, 16, 30, 14]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        wc(ws, 4, c, h,
           font=hf(10, True, DARK_BG), fill_=fill(GOLD),
           align=center(), border=tborder(DARK_BG))
        wcol(ws, c, w)
    ws.row_dimensions[4].height = 20

    # ── Class rows ────────────────────────────────────────────────────────────
    row = 5
    for cls in ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']:
        cnt  = sum(1 for r in results if r['day_classification'] == cls)
        pct  = round(cnt / total * 100, 1) if total else 0
        clr  = CLASS_COLORS[cls]
        lbl  = CLASS_LABELS[cls]
        bg   = RAISED_BG if row % 2 == 0 else CARD_BG

        wc(ws, row, 1, lbl,
           font=Font(name='Arial', size=10, bold=True, color=clr),
           fill_=fill(bg), align=left(), border=tborder())
        wc(ws, row, 2, cnt,
           font=hf(10, True, WHITE), fill_=fill(bg), align=center(), border=tborder())
        wc(ws, row, 3, pct / 100,
           font=cf(10, False, WHITE), fill_=fill(bg), align=center(),
           border=tborder(), fmt='0.0%')

        # Bar column — filled blocks proportional to pct
        bar_len   = int(round(pct / 100 * 25))
        bar_str   = '█' * bar_len + '░' * (25 - bar_len)
        bar_label = f'{bar_str}  {pct}%'
        wc(ws, row, 4, bar_label,
           font=Font(name='Courier New', size=9, bold=False, color=clr),
           fill_=fill(bg), align=left(), border=tborder())

        wc(ws, row, 5, '',
           fill_=fill(clr), align=center(), border=tborder())

        ws.row_dimensions[row].height = 18
        row += 1

    # ── Totals row ────────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 4
    row += 1
    ws.merge_cells(f'A{row}:B{row}')
    wc(ws, row, 1, f'TOTAL  —  {total} days',
       font=hf(10, True, GOLD), fill_=fill(RAISED_BG),
       align=center(), border=tborder())
    wc(ws, row, 3, 1.0,
       font=hf(10, True, GOLD), fill_=fill(RAISED_BG),
       align=center(), border=tborder(), fmt='0.0%')
    for c in [4, 5]:
        wc(ws, row, c, '', fill_=fill(RAISED_BG), border=tborder())
    ws.row_dimensions[row].height = 20

    # ── DOW breakdown ─────────────────────────────────────────────────────────
    row += 2
    ws.merge_cells(f'A{row}:E{row}')
    wc(ws, row, 1, 'DAY-OF-WEEK BREAKDOWN',
       font=hf(11, True, DARK_BG), fill_=fill(GOLD),
       align=center(), border=tborder(DARK_BG))
    ws.row_dimensions[row].height = 22
    row += 1

    dow_headers = ['Day', 'DWP', 'DNP', 'R1', 'R2', 'Unclassified', 'Total']
    dow_colors  = [WHITE, TEAL, GOLD, BLUE_CLR, ORANGE, MUTED, WHITE]
    for c, (h, clr) in enumerate(zip(dow_headers, dow_colors), 1):
        wc(ws, row, c, h,
           font=hf(9, True, DARK_BG), fill_=fill(GOLD if c == 1 else RAISED_BG),
           align=center(), border=tborder(DARK_BG))
    ws.row_dimensions[row].height = 18
    row += 1

    # extend col widths for DOW table
    wcol(ws, 1, 14)
    for c in range(2, 8):
        wcol(ws, c, 13)

    for i, dow in enumerate(DAYS_OF_WEEK[:5]):  # Mon-Fri only
        bg      = RAISED_BG if i % 2 == 0 else CARD_BG
        day_recs = [r for r in results if r['day_of_week'] == dow]
        day_tot  = len(day_recs)
        wc(ws, row, 1, dow, font=hf(9, True, WHITE), fill_=fill(bg),
           align=left(), border=tborder())
        for c_idx, cls in enumerate(['DWP', 'DNP', 'R1', 'R2', 'Unclassified'], 2):
            cnt = sum(1 for r in day_recs if r['day_classification'] == cls)
            clr = CLASS_COLORS[cls]
            wc(ws, row, c_idx, cnt,
               font=Font(name='Arial', size=9, bold=bool(cnt), color=clr if cnt else MUTED),
               fill_=fill(bg), align=center(), border=tborder())
        wc(ws, row, 7, day_tot,
           font=hf(9, True, GOLD), fill_=fill(bg), align=center(), border=tborder())
        ws.row_dimensions[row].height = 16
        row += 1

    ws.freeze_panes = 'A5'


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — CLASSIFICATION SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_class_sheet(wb, cls, rows):
    clr   = CLASS_COLORS[cls]
    label = CLASS_LABELS[cls]

    ws = wb.create_sheet(title=cls)
    ws.sheet_properties.tabColor = clr
    ws.sheet_view.showGridLines  = False

    total = len(rows)

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:D1')
    wc(ws, 1, 1, label,
       font=hf(13, True, DARK_BG), fill_=fill(clr),
       align=center(), border=tborder(DARK_BG))
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:D2')
    wc(ws, 2, 1, f'{total} days classified as {cls}',
       font=cf(10, False, MUTED), fill_=fill(RAISED_BG),
       align=center(), border=tborder())
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6

    # ── Column headers ────────────────────────────────────────────────────────
    col_headers = ['Date', 'Day of Week', 'Classification', 'Reason']
    col_widths  = [14,     14,            18,               64]

    for c, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        wc(ws, 4, c, h,
           font=hf(10, True, DARK_BG), fill_=fill(clr),
           align=center(), border=tborder(DARK_BG))
        wcol(ws, c, w)
    ws.row_dimensions[4].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, rec in enumerate(rows):
        bg      = RAISED_BG if i % 2 == 0 else CARD_BG
        excel_r = i + 5

        wc(ws, excel_r, 1, rec['date'],
           font=cf(10, False, WHITE), fill_=fill(bg),
           align=center(), border=tborder())
        wc(ws, excel_r, 2, rec['day_of_week'],
           font=cf(10, False, MUTED), fill_=fill(bg),
           align=center(), border=tborder())
        wc(ws, excel_r, 3, rec['day_classification'],
           font=Font(name='Arial', size=10, bold=True, color=clr),
           fill_=fill(bg), align=center(), border=tborder())
        wc(ws, excel_r, 4, rec['class_reason'],
           font=cf(9, False, WHITE), fill_=fill(bg),
           align=left(), border=tborder())
        ws.row_dimensions[excel_r].height = 16

    if not rows:
        ws.merge_cells('A5:D5')
        wc(ws, 5, 1, 'No days classified as ' + cls,
           font=cf(10, False, MUTED), fill_=fill(CARD_BG), align=center())
        ws.row_dimensions[5].height = 20

    ws.freeze_panes = 'A5'


# ══════════════════════════════════════════════════════════════════════════════
# MASTER BUILD
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(results, out_dir):
    wb = Workbook()

    build_summary_sheet(wb, results)

    for cls in ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']:
        cls_rows = [r for r in results if r['day_classification'] == cls]
        build_class_sheet(wb, cls, cls_rows)

    ts    = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f'DailyClassification_{ts}.xlsx'
    path  = os.path.join(out_dir, fname)
    wb.save(path)
    log.info(f"Saved: {path}")
    return path


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print()
    print('=' * 60)
    print('  DAILY CLASSIFICATION ENGINE — Standalone')
    print('=' * 60)
    print()
    print('  Day types : DWP | DNP | R1 | R2 | Unclassified')
    print('  Output    : Summary + 5 classification sheets')
    print()

    files = find_csv_files()
    if not files:
        filepath = input('  M1 data CSV path: ').strip().strip('"').strip("'")
    elif len(files) == 1:
        filepath = files[0]
        log.info(f'  Found: {filepath}')
    else:
        print('  Files found:')
        for i, f in enumerate(files):
            print(f'    [{i+1}] {f}')
        while True:
            try:
                c = int(input('  Select: ')) - 1
                if 0 <= c < len(files):
                    filepath = files[c]; break
            except (ValueError, KeyboardInterrupt):
                pass

    df_m1 = load_ohlc(filepath)

    print('\n  Date range filter (leave blank for all data):')
    date_range = None
    try:
        s = input('    Start (YYYY-MM-DD or blank): ').strip()
        e = input('    End   (YYYY-MM-DD or blank): ').strip()
        if s and e:
            from datetime import date
            date_range = (datetime.strptime(s, '%Y-%m-%d').date(),
                          datetime.strptime(e, '%Y-%m-%d').date())
            log.info(f'    Date range: {date_range[0]} -> {date_range[1]}')
        else:
            log.info('    No date range — full dataset')
    except Exception:
        log.info('    Invalid date — full dataset')

    print()
    log.info('Running classification...')
    results = run_classification(df_m1, date_range=date_range)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    path    = build_excel(results, out_dir)

    print()
    print('=' * 60)
    print('  DONE')
    print(f'  Total days : {len(results)}')
    print()
    print('  Classification breakdown:')
    for cls in ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']:
        cnt = sum(1 for r in results if r['day_classification'] == cls)
        pct = round(cnt / len(results) * 100, 1) if results else 0
        print(f'    {cls:>14} : {cnt:>5}  ({pct}%)')
    print(f'\n  Output : {path}')
    print('=' * 60)


if __name__ == '__main__':
    main()
