#!/usr/bin/env python3
"""
FVG Phase 2 — OHLC Grid Search & Leaderboard Builder
=====================================================
Takes Phase 1 output (entry price, direction, entry datetime per trade)
and replays each trade against raw OHLC bars for every SL%/TP% combination.

Grid boundaries are derived from Phase 1 MAE%/MFE% min/max — no arbitrary values.

Stop/target logic (exact match to Phase 1 calc_mfe_pct):
  SHORT: stop hit if bar.high >= stop_price  | target hit if bar.low <= target_price
  LONG:  stop hit if bar.low  <= stop_price  | target hit if bar.high >= target_price
  Stop is checked BEFORE target on the same bar.
  Scan from entry bar to EOD (16:59).

Pre-filters (applied before writing to sheet):
  1. Blown = YES  (equity hit zero at any point)
  2. EV_R < 0.5   (not enough edge)

All formulas sourced directly from TV Backtester template rows.
Output headers are identical to what master_backtester.py expects.
"""

import os, sys, glob, math, logging, argparse
import numpy as np
import pandas as pd
import duckdb
from pathlib import Path
from datetime import datetime, time as dtime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(message)s')
log = logging.getLogger(__name__)

# ── DEFAULTS ──────────────────────────────────────────────────────────────────
DEFAULT_ACCOUNT    = 4500
DEFAULT_RISK       = 225
GRID_STEP          = 0.01   # % step for both SL and TP grid
EOD_HOUR           = 16
EOD_MINUTE         = 59
DEFAULT_MIN_RR        = 0.5           # Minimum RR (TP/SL) — configurable at startup
DEFAULT_MIN_EVR       = None          # Minimum EV_R — None = no filter

# ── STYLING (same palette as Phase 1) ────────────────────────────────────────
DARK_BG    = '0D0F14'
GOLD       = 'F5C842'
TEAL       = '3DD9B3'
RED_CLR    = 'F5504A'
WHITE      = 'E8EAF0'
MUTED      = '7A82A0'
CARD_BG    = '13161E'
RAISED_BG  = '1A1E28'
BORDER_CLR = '252A38'
ORANGE     = 'F5A623'

def fill(c):  return PatternFill('solid', fgColor=c)
def center(): return Alignment(horizontal='center', vertical='center')
def left():   return Alignment(horizontal='left',   vertical='center')
def tborder(c=BORDER_CLR):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def wcol(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def wc(ws, r, c, v, font=None, fill_=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font          = font
    if fill_:  cell.fill          = fill_
    if align:  cell.alignment     = align
    if border: cell.border        = border
    if fmt:    cell.number_format = fmt
    return cell
def hf(sz=11, bold=True,  color=WHITE): return Font(name='Arial', size=sz, bold=bold,  color=color)
def cf(sz=10, bold=False, color=WHITE): return Font(name='Arial', size=sz, bold=bold,  color=color)

# ── ERROR HANDLING ────────────────────────────────────────────────────────────
def abort(msg):
    print(f"\n❌  ERROR\n{'─'*60}\n{msg}\n{'─'*60}")
    sys.exit(1)

def warn(msg):
    print(f"⚠️  WARNING: {msg}")

# ── FILE DETECTION ─────────────────────────────────────────────────────────────
def pick_file(prompt, candidates, label):
    if not candidates:
        abort(
            f"No {label} files found in the script folder.\n"
            f"Please place your {label} file in the same folder as this script."
        )
    if len(candidates) == 1:
        print(f"  ✅  Auto-selected {label}: {os.path.basename(candidates[0])}")
        return candidates[0]

    print(f"\n📂  {label} files found:\n")
    for i, f in enumerate(candidates, 1):
        size_kb = os.path.getsize(f) // 1024
        marker  = " ← Phase 1 detected" if 'FVG_Phase1_' in os.path.basename(f) else ""
        print(f"    [{i}]  {os.path.basename(f)}  ({size_kb:,} KB){marker}")
    print()
    while True:
        try:
            choice = input(f"  👉  {prompt}: ").strip()
            idx    = int(choice) - 1
            if 0 <= idx < len(candidates):
                return candidates[idx]
            print(f"  ❌  Please enter a number between 1 and {len(candidates)}")
        except ValueError:
            print("  ❌  Please enter a valid number")
        except KeyboardInterrupt:
            print("\n  Cancelled.")
            sys.exit(0)

DB_PATH = Path(__file__).parent.parent / 'Fractal Sweep' / 'candle_science.duckdb'


def find_files():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    print("\n" + "="*60)
    print("  FVG PHASE 2 — Grid Search & Leaderboard Builder")
    print("="*60)

    xlsx_files = sorted([
        f for f in glob.glob(os.path.join(script_dir, '*.xlsx'))
        if not f.endswith('_PHASE2_LEADERBOARD.xlsx')
        and not f.endswith('_BACKTESTER_OUTPUT.xlsx')
        and not os.path.basename(f).startswith('~$')
        and not os.path.basename(f).startswith('FVG2_')
    ])
    phase1_file = pick_file("Enter number of your PHASE 1 output file", xlsx_files, "Phase 1 xlsx")

    return phase1_file, script_dir


def load_ohlc(table='nq_1m'):
    """Load all RTH bars from DuckDB, converted to ET."""
    con = duckdb.connect(str(DB_PATH), read_only=True)
    df = con.execute(f"""
        SELECT
            timezone('America/New_York', timestamp) AS datetime,
            open, high, low, close
        FROM {table}
        WHERE date_part('hour', timezone('America/New_York', timestamp)) BETWEEN 9 AND 16
          AND date_part('dow',  timezone('America/New_York', timestamp)) BETWEEN 1 AND 5
        ORDER BY timestamp
    """).df()
    con.close()
    df['datetime'] = pd.to_datetime(df['datetime']).dt.tz_localize(None)
    df.sort_values('datetime', inplace=True)
    df.reset_index(drop=True, inplace=True)
    log.info(f"  OHLC loaded: {len(df):,} bars from {table} | "
             f"{df['datetime'].min().date()} to {df['datetime'].max().date()}")
    return df


# ── LOAD PHASE 1 TRADES (v2) ──────────────────────────────────────────────────
# Reads all classification Data sheets from Phase 1 v2 output.
# Each trade is tagged with its classification.
# Returns:
#   all_trades — flat list of all trades
#   by_class   — dict { 'DWP': [...], 'DNP': [...], ... }
#   mae_vals   — all MAE% values (global grid bounds)
#   mfe_vals   — all MFE% values (global grid bounds)
# ─────────────────────────────────────────────────────────────────────────────
CLASSIFICATIONS    = ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']
DATA_SHEET_SUFFIX  = ' Data'

NEEDED_COLS = {
    'date':        ['date'],
    'trade':       ['trade?', 'trade'],
    'direction':   ['trade dir', 'trade direction', 'direction'],
    'entry_price': ['entry price', 'entry'],
    'fill_time':   ['fill time', 'fill'],
    'mae':         ['mae %', 'mae%', 'mae'],
    'mfe':         ['ext mfe %', 'mfe %', 'mfe%', 'mfe'],
}


def _parse_sheet_trades(ws, classification):
    """Parse all trade rows from one classification Data sheet."""
    hdr_row = None
    col_map = {}
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = {str(ws.cell(r, c).value or '').strip().lower(): c
                    for c in range(1, ws.max_column + 1)}
        found = {}
        for key, aliases in NEEDED_COLS.items():
            for alias in aliases:
                if alias in row_vals:
                    found[key] = row_vals[alias]
                    break
        if len(found) >= 5:
            hdr_row = r
            col_map = found
            break

    if not hdr_row:
        return [], 0

    trades  = []
    skipped = 0

    for r in range(hdr_row + 1, ws.max_row + 1):
        if str(ws.cell(r, col_map['trade']).value or '').strip().upper() != 'YES':
            continue
        try:
            date_val    = ws.cell(r, col_map['date']).value
            direction   = str(ws.cell(r, col_map['direction']).value or '').strip().upper()
            entry_price = float(ws.cell(r, col_map['entry_price']).value)
            fill_raw    = ws.cell(r, col_map['fill_time']).value
            mae         = float(ws.cell(r, col_map['mae']).value)
            mfe         = float(ws.cell(r, col_map['mfe']).value)

            if direction not in ('LONG', 'SHORT'):
                skipped += 1
                continue

            if isinstance(date_val, datetime):
                trade_date = date_val.date()
            else:
                trade_date = pd.to_datetime(str(date_val)).date()

            if isinstance(fill_raw, dtime):
                t = fill_raw
            else:
                parts = [int(x) for x in str(fill_raw).split(':')]
                t = dtime(parts[0], parts[1], parts[2] if len(parts) > 2 else 0)

            entry_dt = datetime.combine(trade_date, t)

            trades.append({
                'date':           trade_date,
                'direction':      direction,
                'entry_price':    entry_price,
                'entry_dt':       entry_dt,
                'mae':            mae,
                'mfe':            mfe,
                'classification': classification,
            })
        except Exception:
            skipped += 1
            continue

    return trades, skipped


def load_phase1_trades(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        abort(f"Could not open Phase 1 file: {os.path.basename(filepath)}\n"
              f"Make sure it is not open in Excel.\nDetail: {str(e)}")

    all_trades    = []
    by_class      = {cls: [] for cls in CLASSIFICATIONS}
    total_skipped = 0

    for cls in CLASSIFICATIONS:
        sheet_name = cls + DATA_SHEET_SUFFIX
        if sheet_name not in wb.sheetnames:
            warn(f"Sheet '{sheet_name}' not found — skipping.")
            continue
        ws = wb[sheet_name]
        trades, skipped = _parse_sheet_trades(ws, cls)
        total_skipped  += skipped
        by_class[cls]   = trades
        all_trades.extend(trades)
        log.info(f"  {cls:>14} Data: {len(trades):>5} trades")

    if not all_trades:
        abort(f"No valid trade rows found in any classification sheet.\n"
              f"Make sure Phase 1 v2 output is used.")

    if total_skipped:
        warn(f"{total_skipped} trade rows skipped due to missing or invalid values.")

    mae_vals = [t['mae'] for t in all_trades]
    mfe_vals = [t['mfe'] for t in all_trades]

    log.info(f"  {'TOTAL':>14}      : {len(all_trades):>5} trades")
    log.info(f"  Global MAE range : {min(mae_vals):.4f}% to {max(mae_vals):.4f}%")
    log.info(f"  Global MFE range : {min(mfe_vals):.4f}% to {max(mfe_vals):.4f}%")

    return all_trades, by_class, mae_vals, mfe_vals


# ── BUILD BAR ARRAYS PER TRADE ────────────────────────────────────────────────
def build_trade_bar_arrays(trades, df_ohlc):
    df_ohlc            = df_ohlc.copy()
    df_ohlc['date']    = df_ohlc['datetime'].dt.date
    df_ohlc['time_s']  = df_ohlc['datetime'].dt.time
    eod_time           = dtime(EOD_HOUR, EOD_MINUTE)

    trade_arrays = []
    missing      = 0

    for t in trades:
        day_bars = df_ohlc[
            (df_ohlc['date'] == t['date']) &
            (df_ohlc['datetime'] >= t['entry_dt']) &
            (df_ohlc['time_s'] <= eod_time)
        ].reset_index(drop=True)

        if len(day_bars) == 0:
            missing += 1
            trade_arrays.append(None)
            continue

        trade_arrays.append({
            'direction':   t['direction'],
            'entry_price': t['entry_price'],
            'highs':       day_bars['high'].values.astype(np.float64),
            'lows':        day_bars['low'].values.astype(np.float64),
        })

    if missing:
        warn(f"{missing} trades had no matching OHLC bars and were skipped.")

    valid = sum(1 for a in trade_arrays if a is not None)
    log.info(f"  Trade bar arrays: {valid} valid of {len(trades)}")
    return trade_arrays


# ── VECTORIZED SIMULATION ENGINE ─────────────────────────────────────────────
# Uses numpy to find first stop/target hit across all bars at once.
# Logic is identical to simulate_trade (Python loop reference).
# Stop is checked BEFORE target — enforced via argmin on stop mask first.
# ─────────────────────────────────────────────────────────────────────────────

def simulate_trade_vec(bars, sl_price, tp_price):
    """
    Vectorized version of simulate_trade.
    Finds the first bar where stop or target is hit using numpy argmax.
    Stop checked before target on same bar — same rule as Python loop version.
    Returns 'W' or 'L'.
    """
    highs     = bars['highs']
    lows      = bars['lows']
    direction = bars['direction']

    if direction == 'SHORT':
        stop_mask   = highs >= sl_price   # stop hit if high >= sl
        target_mask = lows  <= tp_price   # target hit if low <= tp
    else:
        stop_mask   = lows  <= sl_price   # stop hit if low <= sl
        target_mask = highs >= tp_price   # target hit if high >= tp

    stop_hit   = np.any(stop_mask)
    target_hit = np.any(target_mask)

    if not stop_hit and not target_hit:
        return 'L'  # EOD — neither hit

    if stop_hit and not target_hit:
        return 'L'

    if target_hit and not stop_hit:
        return 'W'

    # Both hit — find which came first
    # Stop checked before target on same bar: if same bar, stop wins
    first_stop   = int(np.argmax(stop_mask))
    first_target = int(np.argmax(target_mask))

    return 'L' if first_stop <= first_target else 'W'


def simulate_combo_vec(trade_arrays, sl_pct, tp_pct, account_size, risk_per_trade, sharpe_n=1):
    """
    Vectorized version of simulate_combo.
    Uses simulate_trade_vec instead of simulate_trade.
    All metric calculations are identical.
    """
    outcomes = []
    pnl_list = []
    rr       = tp_pct / sl_pct

    for bars in trade_arrays:
        if bars is None:
            continue
        entry = bars['entry_price']
        dir_  = bars['direction']

        if dir_ == 'SHORT':
            sl_price = entry * (1 + sl_pct / 100)
            tp_price = entry * (1 - tp_pct / 100)
        else:
            sl_price = entry * (1 - sl_pct / 100)
            tp_price = entry * (1 + tp_pct / 100)

        result = simulate_trade_vec(bars, sl_price, tp_price)
        outcomes.append(result)
        pnl_list.append(risk_per_trade * rr if result == 'W' else -risk_per_trade)

    n_total = len(outcomes)
    if n_total == 0:
        return None

    n_wins   = outcomes.count('W')
    n_losses = outcomes.count('L')
    wr       = n_wins   / n_total
    lr       = n_losses / n_total

    win_pnls  = [p for p in pnl_list if p > 0]
    loss_pnls = [abs(p) for p in pnl_list if p < 0]

    avg_win_d  = sum(win_pnls)  / len(win_pnls)  if win_pnls  else 0.0
    avg_loss_d = sum(loss_pnls) / len(loss_pnls) if loss_pnls else risk_per_trade
    gross_w    = sum(win_pnls)
    gross_l    = sum(loss_pnls)
    total_pl   = gross_w - gross_l

    avg_win_r  = avg_win_d / avg_loss_d if avg_loss_d > 0 else 0.0
    total_r    = (n_wins * avg_win_r) - (n_losses * 1.0)
    ev_dollar  = (wr * avg_win_d) - (lr * avg_loss_d)
    ev_r       = ev_dollar / avg_loss_d if avg_loss_d > 0 else 0.0
    pf         = gross_w / gross_l if gross_l > 0 else 999.0
    ce         = ev_r * pf
    risk_pct   = risk_per_trade / account_size
    n_bankroll = int(1.0 / risk_pct) if risk_pct > 0 else 20
    ror_pct    = 100.0 if ce <= 0 else min(100.0, ((1-ce)/(1+ce))**n_bankroll * 100)

    try:
        mcl_calc = math.ceil(math.log(n_total) / math.log(1.0/(1.0-wr))) if 0 < wr < 1 else n_losses
    except (ValueError, ZeroDivisionError):
        mcl_calc = n_losses

    max_dd_pct_calc = mcl_calc * risk_per_trade / account_size
    r_arr  = np.array([p / risk_per_trade for p in pnl_list])
    std_r  = float(np.std(r_arr, ddof=1)) if len(r_arr) > 1 else 0.0
    sqn    = round((ev_r / std_r) * math.sqrt(n_total), 4) if std_r > 0 else 0.0

    # Sharpe — per trade dollar PnL, annualised by unique trading days in this bucket
    # N passed in from caller (unique dates in this classification bucket)
    pnl_arr  = np.array(pnl_list, dtype=np.float64)
    mean_pnl = float(np.mean(pnl_arr))
    std_pnl  = float(np.std(pnl_arr, ddof=1)) if len(pnl_arr) > 1 else 0.0
    sharpe   = round((mean_pnl / std_pnl) * math.sqrt(sharpe_n), 4) if std_pnl > 0 else 0.0

    def max_run(lst, target):
        max_s = cur_s = 0
        for o in lst:
            if o == target: cur_s += 1; max_s = max(max_s, cur_s)
            else: cur_s = 0
        return max_s

    max_w_run = max_run(outcomes, 'W')
    max_l_run = max_run(outcomes, 'L')

    equity   = float(account_size)
    peak     = float(account_size)
    low_eq   = float(account_size)
    max_dd_d = 0.0
    blown    = False
    breaches = 0

    for p in pnl_list:
        equity += p
        if equity > peak: peak = equity
        dd = peak - equity
        if dd > max_dd_d: max_dd_d = dd
        if equity < low_eq: low_eq = equity
        if equity <= 0: blown = True
        if equity < account_size: breaches += 1

    dd_pct = (max_dd_d / peak * 100) if peak > 0 else 0.0
    drr    = max_dd_d / risk_per_trade if risk_per_trade > 0 else 0.0

    return {
        'sl':           round(sl_pct, 4),
        'tp':           round(tp_pct, 4),
        'rr':           round(rr, 4),
        'trades':       n_total,
        'wins':         n_wins,
        'losses':       n_losses,
        'be':           0,
        'win_pct':      round(wr * 100, 4),
        'avg_win':      round(avg_win_d, 4),
        'avg_loss':     round(avg_loss_d, 4),
        'ratio_wl':     round(avg_win_d / avg_loss_d, 4) if avg_loss_d > 0 else 0.0,
        'gross_w':      round(gross_w, 4),
        'gross_l':      round(gross_l, 4),
        'total_pl':     round(total_pl, 4),
        'largest_win':  round(max(win_pnls)  if win_pnls  else 0.0, 4),
        'largest_loss': round(-max(loss_pnls) if loss_pnls else 0.0, 4),
        'ev_dollar':    round(ev_dollar, 4),
        'pf':           round(pf, 4),
        'ce':           round(ce, 6),
        'sqn':          round(sqn, 4),
        'sharpe':       sharpe,
        'max_streak':   max(max_w_run, max_l_run),
        'max_w_run':    max_w_run,
        'max_l_run':    max_l_run,
        'dd_pct':       round(dd_pct, 4),
        'drr':          round(drr, 4),
        'ror_pct':      round(ror_pct, 4),
        'ev_r':         round(ev_r, 6),
        'total_r':      round(total_r, 4),
        'mcl_calc':     mcl_calc,
        'max_dd_d':     round(max_dd_d, 4),
        'low_eq':       round(low_eq, 4),
        'blown':        'YES' if blown else 'NO',
        'breaches':     breaches,
    }


# ── VECTORIZATION AUDIT ───────────────────────────────────────────────────────
# Runs Python loop vs vectorized on a sample of combos and trade arrays.
# Must pass before vectorized engine is used for the full grid search.
# Tolerance: outcomes must be identical (W/L match 100%), metrics within 1e-6.
# ─────────────────────────────────────────────────────────────────────────────
AUDIT_SAMPLE_TRADES  = 50    # max trades to sample for audit
AUDIT_SAMPLE_COMBOS  = 25    # combos to test per audit
AUDIT_METRIC_TOL     = 1e-4  # tolerance for float metric comparison

def run_vectorization_audit(trade_arrays, sl_values, tp_values,
                             account_size, risk_per_trade):
    """
    Compares Python loop simulate_combo vs vectorized simulate_combo_vec
    on a random sample of SL/TP combos using the provided trade arrays.
    Aborts if any outcome or metric mismatch is found.
    Returns True if audit passes.
    """
    import random

    # Sample trade arrays
    valid_arrays = [a for a in trade_arrays if a is not None]
    if not valid_arrays:
        abort("Audit failed: no valid trade arrays to test.")

    sample_arrays = valid_arrays[:AUDIT_SAMPLE_TRADES]

    # Sample combos — pick evenly spaced from the grid
    step = max(1, len(sl_values) * len(tp_values) // AUDIT_SAMPLE_COMBOS)
    all_combos = [(sl, tp) for sl in sl_values for tp in tp_values]
    sample_combos = all_combos[::step][:AUDIT_SAMPLE_COMBOS]

    mismatches = []

    for sl, tp in sample_combos:
        r_loop = simulate_combo(sample_arrays, sl, tp, account_size, risk_per_trade)
        r_vec  = simulate_combo_vec(sample_arrays, sl, tp, account_size, risk_per_trade)

        if r_loop is None and r_vec is None:
            continue
        if r_loop is None or r_vec is None:
            mismatches.append(f"SL={sl} TP={tp}: one returned None")
            continue

        # Check key metrics
        for key in ['wins', 'losses', 'trades', 'blown']:
            if r_loop[key] != r_vec[key]:
                mismatches.append(
                    f"SL={sl} TP={tp}: {key} mismatch "
                    f"loop={r_loop[key]} vec={r_vec[key]}")

        for key in ['win_pct', 'ev_r', 'pf', 'ce', 'ror_pct', 'total_pl', 'sharpe']:
            diff = abs(float(r_loop[key]) - float(r_vec[key]))
            if diff > AUDIT_METRIC_TOL:
                mismatches.append(
                    f"SL={sl} TP={tp}: {key} diff={diff:.8f} "
                    f"loop={r_loop[key]} vec={r_vec[key]}")

    if mismatches:
        msg = "VECTORIZATION AUDIT FAILED — mismatches found:\n"
        for m in mismatches[:10]:
            msg += f"  {m}\n"
        if len(mismatches) > 10:
            msg += f"  ... and {len(mismatches)-10} more"
        abort(msg)

    log.info(f"  Audit PASSED — {len(sample_combos)} combos x "
             f"{len(sample_arrays)} trades — loop vs vec match within {AUDIT_METRIC_TOL}")
    return True


# ── SIMULATE ONE TRADE ────────────────────────────────────────────────────────
def simulate_trade(bars, sl_price, tp_price):
    """
    Bar-by-bar replay. Stop checked BEFORE target on same bar.
    Exact logic from Phase 1 calc_mfe_pct.
    EOD without hitting target = LOSS.
    """
    highs     = bars['highs']
    lows      = bars['lows']
    direction = bars['direction']

    if direction == 'SHORT':
        for i in range(len(highs)):
            if highs[i] >= sl_price: return 'L'   # stop hit
            if lows[i]  <= tp_price: return 'W'   # target hit
    else:  # LONG
        for i in range(len(highs)):
            if lows[i]  <= sl_price: return 'L'   # stop hit
            if highs[i] >= tp_price: return 'W'   # target hit

    return 'L'  # EOD — target never hit

# ── SIMULATE ONE SL/TP COMBO ─────────────────────────────────────────────────
def simulate_combo(trade_arrays, sl_pct, tp_pct, account_size, risk_per_trade, sharpe_n=1):
    outcomes = []
    pnl_list = []
    rr       = tp_pct / sl_pct

    for bars in trade_arrays:
        if bars is None:
            continue

        entry = bars['entry_price']
        dir_  = bars['direction']

        if dir_ == 'SHORT':
            sl_price = entry * (1 + sl_pct / 100)
            tp_price = entry * (1 - tp_pct / 100)
        else:
            sl_price = entry * (1 - sl_pct / 100)
            tp_price = entry * (1 + tp_pct / 100)

        result = simulate_trade(bars, sl_price, tp_price)
        outcomes.append(result)
        pnl_list.append(risk_per_trade * rr if result == 'W' else -risk_per_trade)

    n_total = len(outcomes)
    if n_total == 0:
        return None

    n_wins   = outcomes.count('W')
    n_losses = outcomes.count('L')
    wr       = n_wins  / n_total
    lr       = n_losses / n_total

    win_pnls  = [p for p in pnl_list if p > 0]
    loss_pnls = [abs(p) for p in pnl_list if p < 0]

    avg_win_d  = sum(win_pnls)  / len(win_pnls)  if win_pnls  else 0.0
    avg_loss_d = sum(loss_pnls) / len(loss_pnls) if loss_pnls else risk_per_trade
    gross_w    = sum(win_pnls)
    gross_l    = sum(loss_pnls)
    total_pl   = gross_w - gross_l

    # ── All formulas from TV Backtester template ──────────────────────────────
    # Row 10: Avg Win R = avg_win_$ / avg_loss_$
    avg_win_r  = avg_win_d / avg_loss_d if avg_loss_d > 0 else 0.0
    # Row 11: Avg Loss R = 1.0 always
    # Row 24: Total R = (W × avg_win_R) - (L × avg_loss_R)
    total_r    = (n_wins * avg_win_r) - (n_losses * 1.0)
    # Row 29: EV $ = (win_rate × avg_win_$) - (loss_rate × avg_loss_$)
    ev_dollar  = (wr * avg_win_d) - (lr * avg_loss_d)
    # Row 30: EV R = EV_$ / avg_loss_$
    ev_r       = ev_dollar / avg_loss_d if avg_loss_d > 0 else 0.0
    # Row 32: PF = gross_wins / gross_losses
    pf         = gross_w / gross_l if gross_l > 0 else 999.0
    # Row 33: CE = EV_R × PF
    ce         = ev_r * pf
    # Row 21: N = floor(1 / risk_pct)
    risk_pct   = risk_per_trade / account_size
    n_bankroll = int(1.0 / risk_pct) if risk_pct > 0 else 20
    # Row 50: RoR = ((1-CE)/(1+CE))^N — CE<=0 means 100% ruin
    ror_pct    = 100.0 if ce <= 0 else min(100.0, ((1-ce)/(1+ce))**n_bankroll * 100)
    # Row 51: MCL calc = ceil(ln(N) / ln(1/(1-win_rate)))
    try:
        mcl_calc = math.ceil(math.log(n_total) / math.log(1.0/(1.0-wr))) if 0 < wr < 1 else n_losses
    except (ValueError, ZeroDivisionError):
        mcl_calc = n_losses
    # Row 52: Max DD % = MCL_calc × risk_$ / account
    max_dd_pct_calc = mcl_calc * risk_per_trade / account_size

    # SQN = (EV_R / StdDev_R) × sqrt(N)
    r_arr  = np.array([p / risk_per_trade for p in pnl_list])
    std_r  = float(np.std(r_arr, ddof=1)) if len(r_arr) > 1 else 0.0
    sqn    = round((ev_r / std_r) * math.sqrt(n_total), 4) if std_r > 0 else 0.0

    # Sharpe — per trade dollar PnL, annualised by unique trading days in this bucket
    pnl_arr  = np.array(pnl_list, dtype=np.float64)
    mean_pnl = float(np.mean(pnl_arr))
    std_pnl  = float(np.std(pnl_arr, ddof=1)) if len(pnl_arr) > 1 else 0.0
    sharpe   = round((mean_pnl / std_pnl) * math.sqrt(sharpe_n), 4) if std_pnl > 0 else 0.0

    # Max consecutive streaks from actual outcomes
    def max_run(lst, target):
        max_s = cur_s = 0
        for o in lst:
            if o == target: cur_s += 1; max_s = max(max_s, cur_s)
            else: cur_s = 0
        return max_s

    max_w_run = max_run(outcomes, 'W')
    max_l_run = max_run(outcomes, 'L')

    # Equity curve — for Low Eq $, Max DD $, Blown, Breaches, DRR
    equity   = float(account_size)
    peak     = float(account_size)
    low_eq   = float(account_size)
    max_dd_d = 0.0
    blown    = False
    breaches = 0

    for p in pnl_list:
        equity += p
        if equity > peak: peak = equity
        dd = peak - equity
        if dd > max_dd_d: max_dd_d = dd
        if equity < low_eq: low_eq = equity
        if equity <= 0: blown = True
        if equity < account_size: breaches += 1

    dd_pct = (max_dd_d / peak * 100) if peak > 0 else 0.0
    drr    = max_dd_d / risk_per_trade if risk_per_trade > 0 else 0.0

    return {
        'sl':           round(sl_pct, 4),
        'tp':           round(tp_pct, 4),
        'rr':           round(rr, 4),
        'trades':       n_total,
        'wins':         n_wins,
        'losses':       n_losses,
        'be':           0,
        'win_pct':      round(wr * 100, 4),
        'avg_win':      round(avg_win_d, 4),
        'avg_loss':     round(avg_loss_d, 4),
        'ratio_wl':     round(avg_win_d / avg_loss_d, 4) if avg_loss_d > 0 else 0.0,
        'gross_w':      round(gross_w, 4),
        'gross_l':      round(gross_l, 4),
        'total_pl':     round(total_pl, 4),
        'largest_win':  round(max(win_pnls)  if win_pnls  else 0.0, 4),
        'largest_loss': round(-max(loss_pnls) if loss_pnls else 0.0, 4),
        'ev_dollar':    round(ev_dollar, 4),
        'pf':           round(pf, 4),
        'ce':           round(ce, 6),
        'sqn':          round(sqn, 4),
        'sharpe':       sharpe,
        'max_streak':   max(max_w_run, max_l_run),
        'max_w_run':    max_w_run,
        'max_l_run':    max_l_run,
        'dd_pct':       round(dd_pct, 4),
        'drr':          round(drr, 4),
        'ror_pct':      round(ror_pct, 4),
        'ev_r':         round(ev_r, 6),
        'total_r':      round(total_r, 4),
        'mcl_calc':     mcl_calc,
        'max_dd_d':     round(max_dd_d, 4),
        'low_eq':       round(low_eq, 4),
        'blown':        'YES' if blown else 'NO',
        'breaches':     breaches,
    }

# ── GRADE FUNCTIONS — exact Metadata thresholds ───────────────────────────────
def grade_ev(ev_r):
    if ev_r > 0.40:  return 'A'
    if ev_r >= 0.25: return 'B'
    if ev_r >= 0.10: return 'C'
    if ev_r >= 0.01: return 'D'
    return 'F'

def grade_pf(pf):
    if pf >= 2.0:  return 'A'
    if pf >= 1.6:  return 'B'
    if pf >= 1.4:  return 'C'
    if pf >= 1.2:  return 'D'
    return 'F'

def grade_ce(ce):
    if ce >= 0.7:  return 'A'
    if ce >= 0.4:  return 'B'
    if ce >= 0.2:  return 'C'
    return 'F'

def grade_ror(ror_pct):
    if ror_pct < 1.0:   return 'A'
    if ror_pct <= 5.0:  return 'B'
    if ror_pct <= 10.0: return 'C'
    if ror_pct <= 20.0: return 'D'
    return 'F'

def grade_mcl(mcl):
    if mcl < 7:   return 'A'
    if mcl <= 10: return 'B'
    if mcl <= 12: return 'C'
    if mcl <= 15: return 'D'
    return 'F'

# ── BUILD GLOBAL GRID (from all trades MAE/MFE) ───────────────────────────────
def build_global_grid(mae_vals, mfe_vals, sl_pct=95, tp_pct=95):
    """
    Derives SL/TP grid boundaries from ALL trades combined.
    Caps at sl_pct/tp_pct percentile to avoid outliers exploding the grid.
    Same grid is used for every classification bucket so results are comparable.
    """
    sl_max = round(float(np.percentile(mae_vals, sl_pct)), 4)
    tp_max = round(float(np.percentile(mfe_vals, tp_pct)), 4)
    sl_min = round(max(GRID_STEP, min(mae_vals)), 4)
    tp_min = round(max(GRID_STEP, min(mfe_vals)), 4)

    sl_values = [round(v, 4) for v in np.arange(sl_min, sl_max + GRID_STEP / 2, GRID_STEP)]
    tp_values = [round(v, 4) for v in np.arange(tp_min, tp_max + GRID_STEP / 2, GRID_STEP)]

    log.info(f"  Grid cap: SL {sl_pct}th pct = {sl_max:.4f}%  |  TP {tp_pct}th pct = {tp_max:.4f}%")
    return sl_values, tp_values


# ── RUN GRID SEARCH FOR ONE BUCKET ───────────────────────────────────────────
def run_grid_search(trade_arrays, sl_values, tp_values,
                    account_size, risk_per_trade,
                    label='ALL', min_rr=0.5, min_evr=None, sharpe_n=1):
    """
    Runs the full SL/TP grid against a specific set of trade arrays.
    Uses vectorized simulate_combo_vec (pre-audited).
    label: classification name for logging.
    sharpe_n: unique trading days in this bucket — used as Sharpe annualization N.
    """
    total    = len(sl_values) * len(tp_values)
    n_valid  = sum(1 for a in trade_arrays if a is not None)

    log.info(f"\n  [{label}] {n_valid} trades | "
             f"{len(sl_values)} SL steps x {len(tp_values)} TP steps = {total:,} combos")

    if n_valid == 0:
        log.info(f"  [{label}] No valid trades — skipping.")
        return []

    results  = []
    filtered = 0
    done     = 0
    n_sl     = len(sl_values)

    for sl_idx, sl in enumerate(sl_values):
        for tp in tp_values:
            r = simulate_combo_vec(trade_arrays, sl, tp, account_size, risk_per_trade, sharpe_n=sharpe_n)
            done += 1

            if r is None:
                filtered += 1
                continue
            if r['blown'] == 'YES':
                filtered += 1
                continue
            if r['rr'] < min_rr:
                filtered += 1
                continue
            if min_evr is not None and r['ev_r'] < min_evr:
                filtered += 1
                continue

            r['g_ev']  = grade_ev(r['ev_r'])
            r['g_pf']  = grade_pf(r['pf'])
            r['g_ce']  = grade_ce(r['ce'])
            r['g_ror'] = grade_ror(r['ror_pct'])
            r['g_mcl'] = grade_mcl(r['mcl_calc'])
            results.append(r)

        pct_done = (sl_idx + 1) / n_sl * 100
        print(f"\r  [{label}] {pct_done:5.1f}%  ({sl_idx+1}/{n_sl} SL steps, {done:,}/{total:,} combos)", end='', flush=True)


    print()  # newline after progress bar
    results.sort(key=lambda x: x['ce'], reverse=True)
    evr_str = f"EV_R<{min_evr}" if min_evr is not None else "no EV_R filter"
    log.info(f"  [{label}] Done: {total:,} tested | "
             f"{filtered:,} filtered | {len(results):,} passing | "
             f"Filters: Blown | RR<{min_rr} | {evr_str}")
    return results

# ── CLASSIFICATION COLORS (matches Phase 1 v2) ───────────────────────────────
CLS_COLORS = {
    'DWP':          TEAL,
    'DNP':          GOLD,
    'R1':           '5B9CF6',
    'R2':           ORANGE,
    'Unclassified': MUTED,
    'Combined':     GOLD,
}


# ── WRITE ONE LEADERBOARD SHEET ───────────────────────────────────────────────
def _write_leaderboard(wb, sheet_title, tab_color, results, trades_subset,
                       mae_vals_subset, mfe_vals_subset,
                       sl_values, tp_values,
                       source_phase1, source_ohlc,
                       account_size, risk_per_trade,
                       retest_cutoff, min_rr, min_evr,
                       is_first=False):
    """
    Writes one Leaderboard + Distribution sheet pair for a given bucket.
    is_first=True uses wb.active instead of create_sheet.
    """
    HEADERS = [
        'Rank', 'SL %', 'TP %', 'RR', 'Trades', 'Wins', 'Losses', 'BE',
        'Win %', 'Avg Win $', 'Avg Loss $', 'Ratio W/L',
        'Gross Profit $', 'Gross Loss $', 'Total P&L $',
        'Largest Win $', 'Largest Loss $', 'EV $', 'PF', 'Comb Edge',
        'SQN', 'Sharpe', 'Max Streak', 'Max W Run', 'Max L Run',
        'DD %', 'DRR', 'RoR %',
        'Avg MAE pts', 'Avg MFE pts', 'Max DD $', 'Low Eq $', 'Blown', 'Breaches'
    ]
    COL_W = [
        7,8,8,7,8,7,8,6,
        8,11,11,11,
        14,13,13,
        14,14,10,8,11,
        8,9,11,11,11,
        8,8,8,
        12,12,12,12,8,10
    ]
    NCOLS = len(HEADERS)

    ws = wb.active if is_first else wb.create_sheet(sheet_title)
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color

    n_trades  = len(trades_subset)
    n_passing = len(results)

    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    wc(ws, 1, 1,
       f'LEADERBOARD — {sheet_title} | {n_passing:,} passing combos | Sorted by Combined Edge',
       font=hf(13, True, tab_color), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
    wc(ws, 2, 1,
       f'Trades: {n_trades} | SL: {min(sl_values):.2f}%-{max(sl_values):.2f}% '
       f'TP: {min(tp_values):.2f}%-{max(tp_values):.2f}% | '
       f'Account: ${account_size:,} Risk: ${risk_per_trade:,} | '
       f'Filters: Blown=removed | RR>={min_rr} | '
       f'{("EV_R>=" + str(min_evr)) if min_evr is not None else "EV_R=no filter"} | '
       f'Entry cutoff: {retest_cutoff.strftime("%H:%M") if retest_cutoff else "10:00"} | '
       f'{datetime.now().strftime("%Y-%m-%d %H:%M")}',
       font=cf(10, False, MUTED), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[2].height = 18

    for c, (h, w) in enumerate(zip(HEADERS, COL_W), 1):
        wc(ws, 3, c, h, font=hf(9, True, DARK_BG), fill_=fill(tab_color),
           align=center(), border=tborder(DARK_BG))
        wcol(ws, c, w)
    ws.row_dimensions[3].height = 20

    if not results:
        ws.merge_cells(f'A4:{get_column_letter(NCOLS)}4')
        wc(ws, 4, 1, 'No combos passed filters for this classification.',
           font=cf(10, False, MUTED), fill_=fill(DARK_BG), align=center())
        ws.freeze_panes = 'B4'
        return ws

    avg_mae = round(float(np.mean(mae_vals_subset)), 4) if mae_vals_subset else 0
    avg_mfe = round(float(np.mean(mfe_vals_subset)), 4) if mfe_vals_subset else 0

    _bdr = tborder()
    _aln = center()

    for rank, r in enumerate(results, 1):
        er = rank + 3
        bg = RAISED_BG if rank % 2 == 0 else CARD_BG

        def cell(col, val, fc=WHITE, bold=False, fmt=None, _er=er, _bg=bg):
            c = ws.cell(row=_er, column=col, value=val)
            c.font      = Font(name='Arial', size=9, color=fc, bold=bold)
            c.fill      = PatternFill('solid', fgColor=_bg)
            c.alignment = _aln
            c.border    = _bdr
            if fmt: c.number_format = fmt

        cell(1,  rank,              fc=GOLD if rank<=10 else MUTED, bold=rank<=10)
        cell(2,  r['sl'],            fmt='0.00')
        cell(3,  r['tp'],            fmt='0.00')
        cell(4,  r['rr'],            fmt='0.00')
        cell(5,  r['trades'])
        cell(6,  r['wins'],          fc=TEAL,    bold=True)
        cell(7,  r['losses'],        fc=RED_CLR, bold=True)
        cell(8,  r['be'])
        cell(9,  r['win_pct'],       fc=TEAL if r['win_pct']>=50 else RED_CLR, fmt='0.00')
        cell(10, r['avg_win'],       fmt='#,##0.00')
        cell(11, r['avg_loss'],      fmt='#,##0.00')
        cell(12, r['ratio_wl'],      fmt='0.000')
        cell(13, r['gross_w'],       fc=TEAL,    fmt='#,##0.00')
        cell(14, r['gross_l'],       fc=RED_CLR, fmt='#,##0.00')
        cell(15, r['total_pl'],      fc=TEAL if r['total_pl']>0 else RED_CLR, bold=True, fmt='#,##0.00')
        cell(16, r['largest_win'],   fmt='#,##0.00')
        cell(17, r['largest_loss'],  fmt='#,##0.00')
        cell(18, r['ev_dollar'],     fc=TEAL if r['ev_dollar']>0 else RED_CLR, fmt='#,##0.00')
        cell(19, r['pf'],            fc=TEAL if r['pf']>=1.5 else (ORANGE if r['pf']>=1.0 else RED_CLR), bold=True, fmt='0.000')
        cell(20, r['ce'],            fc=TEAL if r['ce']>0 else RED_CLR, bold=True, fmt='0.000000')
        cell(21, r['sqn'],           fmt='0.0000')
        cell(22, r['sharpe'],        fc=TEAL if r['sharpe']>=1.0 else (ORANGE if r['sharpe']>=0.5 else RED_CLR), bold=True, fmt='0.0000')
        cell(23, r['max_streak'])
        cell(24, r['max_w_run'],     fc=TEAL)
        cell(25, r['max_l_run'],     fc=RED_CLR if r['max_l_run']>=15 else WHITE)
        cell(26, r['dd_pct'],        fmt='0.00')
        cell(27, r['drr'],           fmt='0.00')
        cell(28, r['ror_pct'],       fc=TEAL if r['ror_pct']<1 else (ORANGE if r['ror_pct']<10 else RED_CLR), fmt='0.0000')
        cell(29, avg_mae,             fmt='0.0000')
        cell(30, avg_mfe,             fmt='0.0000')
        cell(31, r['max_dd_d'],      fmt='#,##0.00')
        cell(32, r['low_eq'],        fc=TEAL if r['low_eq']>=account_size*0.5 else RED_CLR, fmt='#,##0.00')
        cell(33, r['blown'],         fc=RED_CLR if r['blown']=='YES' else MUTED, bold=r['blown']=='YES')
        cell(34, r['breaches'])
        ws.row_dimensions[er].height = 14

    ws.freeze_panes = 'B4'
    ws.auto_filter.ref = f'A3:{get_column_letter(NCOLS)}3'
    return ws


# ── WRITE ONE DISTRIBUTION SHEET ──────────────────────────────────────────────
def _write_distribution(wb, sheet_title, tab_color, results,
                         trades_subset, mae_vals_subset, mfe_vals_subset,
                         sl_values, tp_values,
                         source_phase1, source_ohlc,
                         account_size, risk_per_trade,
                         retest_cutoff, min_rr, min_evr):
    ws2 = wb.create_sheet(sheet_title)
    ws2.sheet_view.showGridLines  = False
    ws2.sheet_properties.tabColor = tab_color

    ws2.merge_cells('A1:B1')
    wc(ws2, 1, 1, f'MAE / MFE DISTRIBUTION — {sheet_title}',
       font=hf(12, True, tab_color), fill_=fill(DARK_BG), align=center())
    ws2.row_dimensions[1].height = 28
    wcol(ws2, 1, 34); wcol(ws2, 2, 20)

    wc(ws2, 2, 1, 'Metric', font=hf(9, True, DARK_BG), fill_=fill(tab_color),
       align=center(), border=tborder(DARK_BG))
    wc(ws2, 2, 2, 'Value',  font=hf(9, True, DARK_BG), fill_=fill(tab_color),
       align=center(), border=tborder(DARK_BG))
    ws2.row_dimensions[2].height = 18

    _bdr = tborder(); _aln_l = left(); _aln_c = center()

    def stat(r, label, val, section=False, gold=False, fmt=None):
        bg  = '1C2333' if section else (RAISED_BG if r % 2 == 0 else CARD_BG)
        lfc = GOLD if section else MUTED
        vfc = GOLD if (gold or section) else WHITE
        c1 = ws2.cell(row=r, column=1, value=label)
        c1.font = Font(name='Arial', size=9, bold=section, color=lfc)
        c1.fill = PatternFill('solid', fgColor=bg)
        c1.alignment = _aln_l; c1.border = _bdr
        c2 = ws2.cell(row=r, column=2, value=val)
        c2.font = Font(name='Arial', size=9, bold=gold or section, color=vfc)
        c2.fill = PatternFill('solid', fgColor=bg)
        c2.alignment = _aln_c; c2.border = _bdr
        if fmt: c2.number_format = fmt
        ws2.row_dimensions[r].height = 15

    row = 3
    stat(row, '-- SOURCE FILES --', '', section=True);                 row += 1
    stat(row, 'Phase 1 file',  os.path.basename(source_phase1));       row += 1
    stat(row, 'OHLC file',     os.path.basename(source_ohlc));         row += 1
    stat(row, 'Trades (bucket)', len(trades_subset));                   row += 1
    stat(row, 'Combos tested',  len(sl_values) * len(tp_values));       row += 1
    stat(row, 'Passing combos', len(results));                          row += 1
    row += 1
    stat(row, '-- FILTER SETTINGS --', '', section=True);              row += 1
    stat(row, 'Blown accounts', 'Always removed — hard filter');        row += 1
    stat(row, 'Min RR (TP/SL)', f'>= {min_rr}');                       row += 1
    stat(row, 'Min EV_R', f'>= {min_evr}' if min_evr is not None else 'No filter'); row += 1
    if retest_cutoff:
        stat(row, 'Entry cutoff', retest_cutoff.strftime('%H:%M') + ' (FVG fixed 09:59)'); row += 1

    if mae_vals_subset:
        row += 1
        stat(row, '-- MAE % (SL grid — GLOBAL bounds) --', '', section=True); row += 1
        stat(row, 'Min MAE %',    round(min(mae_vals_subset), 6), gold=True, fmt='0.000000'); row += 1
        stat(row, 'Avg MAE %',    round(float(np.mean(mae_vals_subset)), 6), gold=True, fmt='0.000000'); row += 1
        stat(row, 'Median MAE %', round(float(np.median(mae_vals_subset)), 6), fmt='0.000000'); row += 1
        stat(row, 'Max MAE %',    round(max(mae_vals_subset), 6), gold=True, fmt='0.000000'); row += 1
        for p in [50, 70, 85, 90, 95]:
            stat(row, f'MAE {p}th %ile', round(float(np.percentile(mae_vals_subset, p)), 6), fmt='0.000000'); row += 1

        row += 1
        stat(row, '-- MFE % (TP grid — GLOBAL bounds) --', '', section=True); row += 1
        stat(row, 'Min MFE %',    round(min(mfe_vals_subset), 6), gold=True, fmt='0.000000'); row += 1
        stat(row, 'Avg MFE %',    round(float(np.mean(mfe_vals_subset)), 6), gold=True, fmt='0.000000'); row += 1
        stat(row, 'Median MFE %', round(float(np.median(mfe_vals_subset)), 6), fmt='0.000000'); row += 1
        stat(row, 'Max MFE %',    round(max(mfe_vals_subset), 6), gold=True, fmt='0.000000'); row += 1
        for p in [50, 70, 85, 90, 95]:
            stat(row, f'MFE {p}th %ile', round(float(np.percentile(mfe_vals_subset, p)), 6), fmt='0.000000'); row += 1

        row += 1
        e_ratio = round(float(np.mean(mfe_vals_subset)) / float(np.mean(mae_vals_subset)), 4)                   if float(np.mean(mae_vals_subset)) > 0 else 0
        stat(row, '-- E-RATIO --', '', section=True);                   row += 1
        stat(row, 'E-Ratio (Avg MFE / Avg MAE)', e_ratio, gold=True);   row += 1

    row += 1
    stat(row, '-- FORMULA REFERENCE --', '', section=True);             row += 1
    stat(row, 'Avg Win R',  'avg_win_$ / avg_loss_$  [Row 10]');         row += 1
    stat(row, 'Avg Loss R', '= 1.0 always  [Row 11]');                   row += 1
    stat(row, 'EV $',       '(WR x avg_win_$) - (LR x avg_loss_$)  [Row 29]'); row += 1
    stat(row, 'EV R',       'EV_$ / avg_loss_$  [Row 30]');              row += 1
    stat(row, 'PF',         'Gross Wins / Gross Losses  [Row 32]');       row += 1
    stat(row, 'CE',         'EV_R x PF  [Row 33]');                       row += 1
    stat(row, 'RoR',        '((1-CE)/(1+CE))^N  [Row 50]');              row += 1
    stat(row, 'MCL calc',   'ceil(ln(N)/ln(1/(1-WR)))  [Row 51]');       row += 1

    ws2.freeze_panes = 'A3'


# ── MASTER BUILD EXCEL ────────────────────────────────────────────────────────
def build_excel(results_by_label, trades_by_label, mae_by_label, mfe_by_label,
                sl_values, tp_values,
                source_phase1, source_ohlc, out_dir,
                account_size, risk_per_trade,
                retest_cutoff=None, min_rr=0.5, min_evr=None):
    """
    results_by_label  — dict { label: [combo_results] }
    trades_by_label   — dict { label: [trades] }
    mae_by_label      — dict { label: [mae_vals] }
    mfe_by_label      — dict { label: [mfe_vals] }
    Label order: Combined, DWP, DNP, R1, R2, Unclassified
    """
    wb     = Workbook()
    labels = ['Combined'] + CLASSIFICATIONS  # Combined first

    first  = True
    for label in labels:
        results        = results_by_label.get(label, [])
        trades_subset  = trades_by_label.get(label, [])
        mae_subset     = mae_by_label.get(label, [])
        mfe_subset     = mfe_by_label.get(label, [])
        clr            = CLS_COLORS.get(label, GOLD)

        lb_title   = f'{label} Leaderboard'
        dist_title = f'{label} Distribution'

        _write_leaderboard(wb, lb_title, clr, results, trades_subset,
                           mae_subset, mfe_subset,
                           sl_values, tp_values,
                           source_phase1, source_ohlc,
                           account_size, risk_per_trade,
                           retest_cutoff, min_rr, min_evr,
                           is_first=first)
        first = False

        _write_distribution(wb, dist_title, clr, results,
                            trades_subset, mae_subset, mfe_subset,
                            sl_values, tp_values,
                            source_phase1, source_ohlc,
                            account_size, risk_per_trade,
                            retest_cutoff, min_rr, min_evr)

    base    = os.path.splitext(os.path.basename(source_phase1))[0]
    ts      = datetime.now().strftime('%Y%m%d_%H%M%S')
    import random, string
    uid     = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    ts_short = datetime.now().strftime('%m%d_%H%M')
    outname = f'FVG2_{ts_short}_{uid}.xlsx'
    outpath = os.path.join(out_dir, outname)
    wb.save(outpath)
    return outpath


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--table',   default='nq_1m',         help='DuckDB table (nq_1m or es_1m)')
    parser.add_argument('--account', default=DEFAULT_ACCOUNT, type=float, help='Account size $')
    parser.add_argument('--risk',    default=DEFAULT_RISK,    type=float, help='Risk per trade $')
    parser.add_argument('--min-rr',  default=DEFAULT_MIN_RR,  type=float, help='Minimum RR filter')
    parser.add_argument('--min-evr', default=None,            type=float, help='Minimum EV_R filter')
    parser.add_argument('--sl-pct',  default=95,              type=float, help='SL grid cap percentile (default 95)')
    parser.add_argument('--tp-pct',  default=95,              type=float, help='TP grid cap percentile (default 95)')
    args, _ = parser.parse_known_args()

    account_size   = args.account
    risk_per_trade = args.risk
    min_rr         = args.min_rr
    min_evr        = args.min_evr
    retest_cutoff  = None

    phase1_file, script_dir = find_files()

    pct = risk_per_trade / account_size * 100
    print(f'\n  Account : ${account_size:,}  |  Risk/trade: ${risk_per_trade:,} ({pct:.1f}%)')
    print(f'  Min RR  : {min_rr}  |  Min EV_R: {min_evr if min_evr is not None else "none"}')
    print(f'  Phase 1 : {os.path.basename(phase1_file)}')
    print(f'  OHLC    : {args.table} (DuckDB)')

    # ── Load trades ───────────────────────────────────────────────────────────
    print('\nLoading Phase 1 trades...')
    all_trades, by_class, mae_vals_all, mfe_vals_all = load_phase1_trades(phase1_file)

    # ── Load OHLC ─────────────────────────────────────────────────────────────
    print('\nLoading OHLC data from DuckDB...')
    df_ohlc = load_ohlc(args.table)

    # ── Build global grid (from ALL trades) ───────────────────────────────────
    sl_values, tp_values = build_global_grid(mae_vals_all, mfe_vals_all,
                                              sl_pct=args.sl_pct, tp_pct=args.tp_pct)
    total_combos = len(sl_values) * len(tp_values)
    log.info(f'\n  Global grid: SL {min(sl_values):.4f}%-{max(sl_values):.4f}% | ')
    log.info(f'  TP {min(tp_values):.4f}%-{max(tp_values):.4f}% | {total_combos:,} combos')

    # ── Build trade bar arrays for all buckets ────────────────────────────────
    print('\nBuilding trade bar arrays...')
    arrays_combined = build_trade_bar_arrays(all_trades, df_ohlc)
    arrays_by_class = {}
    for cls in CLASSIFICATIONS:
        arrays_by_class[cls] = build_trade_bar_arrays(by_class[cls], df_ohlc)

    # ── Run vectorization audit (Combined arrays) ─────────────────────────────
    print('\nRunning vectorization audit...')
    run_vectorization_audit(arrays_combined, sl_values, tp_values,
                             account_size, risk_per_trade)
    print('  Vectorization audit PASSED — using vectorized engine.')

    # ── Run grid search for each bucket ───────────────────────────────────────
    print('\nRunning grid search...')
    results_by_label = {}
    trades_by_label  = {}
    mae_by_label     = {}
    mfe_by_label     = {}

    # Combined
    combined_sharpe_n = max(len(set(t['date'] for t in all_trades)), 1)
    results_by_label['Combined'] = run_grid_search(
        arrays_combined, sl_values, tp_values,
        account_size, risk_per_trade,
        label='Combined', min_rr=min_rr, min_evr=min_evr,
        sharpe_n=combined_sharpe_n)
    trades_by_label['Combined'] = all_trades
    mae_by_label['Combined']    = mae_vals_all
    mfe_by_label['Combined']    = mfe_vals_all

    # Per classification
    for cls in CLASSIFICATIONS:
        cls_trades   = by_class[cls]
        cls_arrays   = arrays_by_class[cls]
        cls_mae_vals = [t['mae'] for t in cls_trades]
        cls_mfe_vals = [t['mfe'] for t in cls_trades]
        cls_sharpe_n = max(len(set(t['date'] for t in cls_trades)), 1)

        results_by_label[cls] = run_grid_search(
            cls_arrays, sl_values, tp_values,
            account_size, risk_per_trade,
            label=cls, min_rr=min_rr, min_evr=min_evr,
            sharpe_n=cls_sharpe_n)
        trades_by_label[cls] = cls_trades
        mae_by_label[cls]    = cls_mae_vals
        mfe_by_label[cls]    = cls_mfe_vals

    # ── Build Excel ───────────────────────────────────────────────────────────
    print('\nBuilding Excel output...')
    outpath = build_excel(
        results_by_label, trades_by_label, mae_by_label, mfe_by_label,
        sl_values, tp_values,
        phase1_file, args.table, script_dir,
        account_size, risk_per_trade,
        retest_cutoff=retest_cutoff, min_rr=min_rr, min_evr=min_evr)

    # ── Summary ───────────────────────────────────────────────────────────────
    print(f'\n{"="*60}')
    print(f'  DONE')
    print(f'  Total trades   : {len(all_trades)}')
    print(f'  Combos tested  : {total_combos:,}')
    print(f'  Output         : {os.path.basename(outpath)}')
    print()
    for label in ['Combined'] + CLASSIFICATIONS:
        r = results_by_label.get(label, [])
        n = len(trades_by_label.get(label, []))
        print(f'  {label:>14} : {n:>4} trades | {len(r):>5} passing combos', end='')
        if r:
            top = r[0]
            print(f'  | #1 CE={top["ce"]}  EV_R={top["ev_r"]}  WR={top["win_pct"]}%  Sharpe={top["sharpe"]}')
        else:
            print(f'  | no combos passed')
    print(f'{"="*60}\n')


if __name__ == '__main__':
    main()
