#!/usr/bin/env python3
"""
FVG Phase 1 v2 — Setup Detector & Raw MAE/MFE Collector
=========================================================
Same as v1 plus:
  - Daily Classification Engine (DWP / DNP / R1 / R2 / Unclassified)
  - Output split into 5 Data + 5 Stats sheets — one pair per classification
  - Half days (< 330 RTH M1 bars) → Unclassified
  - Summary sheet: cross-classification overview table

Classification Rules (from KB02 New Bootcamp Transcripts):
  R1  — 4 or more RTH hours wick into the 9:30 M1 candle range
  R2  — NY1 (hours 10+11) stacks away from 9:30 creating a thigh gap,
          then NY2 (hours 12-14) reverts ALL THE WAY back to touch 9:30
  DWP — NY1 leaves 9:30 cleanly (thigh gap); one NY2 hour takes out a
          prior opposing hourly extreme but never reaches 9:30
  DNP — NY1 leaves 9:30 cleanly; NY2 stacks same direction; no opposing
          hourly extreme taken; 9:30 never touched again
  Unclassified — half day, no 9:30 bar, mixed NY1, or criteria not clearly met
"""

import os, sys, logging
import numpy as np
import pandas as pd
import duckdb
from pathlib import Path
from datetime import datetime, time as dtime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Import classification from daily_classifier.py ────────────────────────────
sys.path.insert(0, os.path.dirname(__file__))
from daily_classifier import (
    classify_day,
    get_rth_bars, get_930_candle, get_hourly_extremes, hour_touches_930,
    HALF_DAY_BAR_THRESHOLD, RTH_START, RTH_END,
)

logging.basicConfig(level=logging.INFO, format='%(message)s')
log = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
TICK          = 0.25
EOD_HOUR      = 16
EOD_MINUTE    = 59
ENTRY_HOUR    = 9
ENTRY_MIN     = 31
SCAN_END_HOUR = 9
SCAN_END_MIN  = 59
MIN_RISK_PTS  = 0.25
TP_BPS        = 0.001   # 10 bps = 0.10%

# ── Color palette ──────────────────────────────────────────────────────────────
DARK_BG    = '0D0F14'
GOLD       = 'F5C842'
TEAL       = '3DD9B3'
RED_CLR    = 'F5504A'
BLUE_CLR   = '5B9CF6'
GREEN_CLR  = '2ECC71'
WHITE      = 'E8EAF0'
MUTED      = '7A82A0'
CARD_BG    = '13161E'
RAISED_BG  = '1A1E28'
BORDER_CLR = '252A38'
ORANGE     = 'F5A623'

CLASS_COLORS = {
    'DWP':          TEAL,
    'DNP':          GOLD,
    'R1':           BLUE_CLR,
    'R2':           ORANGE,
    'Unclassified': MUTED,
}

CLASS_LABELS = {
    'DWP':          '📈  DWP — Directional With Pullback',
    'DNP':          '🚀  DNP — Directional No Pullback',
    'R1':           '🔄  R1  — Range 1',
    'R2':           '↩   R2  — Range 2',
    'Unclassified': '❓  Unclassified',
}

# ── Style helpers ──────────────────────────────────────────────────────────────
def fill(c):    return PatternFill('solid', start_color=c, fgColor=c)
def center():   return Alignment(horizontal='center', vertical='center')
def left():     return Alignment(horizontal='left',   vertical='center')
def tborder(c=BORDER_CLR):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def wcol(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def wc(ws, r, c, v, font=None, fill_=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font          = font
    if fill_:  cell.fill          = fill_
    if align:  cell.alignment     = align
    if border: cell.border        = border
    if fmt:    cell.number_format = fmt
    return cell
def hf(sz=11, bold=True,  color=WHITE): return Font(name='Arial', size=sz, bold=bold,  color=color)
def cf(sz=10, bold=False, color=WHITE): return Font(name='Arial', size=sz, bold=bold, color=color)
def _class_color(cls): return CLASS_COLORS.get(cls, MUTED)


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
DB_PATH = Path(__file__).parent.parent / 'Fractal Sweep' / 'candle_science.duckdb'


def load_ohlc(table='nq_1m'):
    """Load all RTH bars from DuckDB, converted to ET. Returns datetime/open/high/low/close."""
    con = duckdb.connect(str(DB_PATH), read_only=True)
    df = con.execute(f"""
        SELECT
            timezone('America/New_York', timestamp) AS datetime,
            open, high, low, close
        FROM {table}
        WHERE date_part('hour', timezone('America/New_York', timestamp)) BETWEEN 9 AND 16
          AND date_part('dow',  timezone('America/New_York', timestamp)) BETWEEN 1 AND 5
        ORDER BY timestamp
    """).df()
    con.close()
    df['datetime'] = pd.to_datetime(df['datetime']).dt.tz_localize(None)
    df.sort_values('datetime', inplace=True)
    df.reset_index(drop=True, inplace=True)
    log.info(f"Loaded {len(df):,} bars from {table} | "
             f"{df['datetime'].min().date()} -> {df['datetime'].max().date()}")
    return df


# ══════════════════════════════════════════════════════════════════════════════
# FVG DETECTION
# ══════════════════════════════════════════════════════════════════════════════
def check_fvg(c1, c2, c3):
    """
    Returns (direction, zone_low, zone_high, stop) or None.
    C2 body filter: bullish requires green C2, bearish requires red C2.
    """
    # Bullish: C3.low > C1.high, green C2
    if c3['low'] > c1['high'] and c2['close'] > c2['open']:
        return ('BULLISH', c1['high'], c3['low'], min(c1['low'], c2['low']))
    # Bearish: C3.high < C1.low, red C2
    if c3['high'] < c1['low'] and c2['close'] < c2['open']:
        return ('BEARISH', c3['high'], c1['low'], max(c1['high'], c2['high']))
    return None


def find_first_fvg(day_bars):
    """
    Scan all bars where C2 falls in [9:31, 9:59].
    Returns first valid FVG with min risk ≥ MIN_RISK_PTS, or None.
    """
    scan_start = dtime(ENTRY_HOUR, ENTRY_MIN)
    scan_end   = dtime(SCAN_END_HOUR, SCAN_END_MIN)
    n = len(day_bars)
    for i in range(n - 2):
        c2 = day_bars[i + 1]
        c2_t = c2['datetime'].time()
        if c2_t < scan_start or c2_t > scan_end:
            continue
        c1 = day_bars[i]; c3 = day_bars[i + 2]
        result = check_fvg(c1, c2, c3)
        if result is None:
            continue
        d, zl, zh, stop = result
        # Long entry at C3.low (= zone_high); Short entry at C3.high (= zone_low)
        entry    = zh if d == 'BULLISH' else zl
        risk_pts = abs(entry - stop)
        if risk_pts < MIN_RISK_PTS:
            continue
        return {
            'direction':   d,
            'zone_low':    zl,
            'zone_high':   zh,
            'stop':        stop,
            'entry':       entry,
            'trade_dir':   'LONG' if d == 'BULLISH' else 'SHORT',
            'risk_pts':    risk_pts,
            'c1_dt':       c1['datetime'],
            'c3_dt':       c3['datetime'],
            'c3_idx':      i + 2,
            'fvg_size_pct': round((zh - zl) / zl * 100, 6),
        }
    return None


# ══════════════════════════════════════════════════════════════════════════════
# FILL + CASHFLOW / EXTENDED OUTCOMES
# ══════════════════════════════════════════════════════════════════════════════
def calc_fill_and_outcomes(fvg, grp_df):
    """
    Scan fill window (C3+1 bar through 16:00 ET) for a fill at entry price.
    If filled, compute cashflow and extended model outcomes.

    Cashflow model  : TP at entry ± TP_BPS; R = 1.0 on win, -1.0 on stop,
                      None if expired at EOD without either trigger.
    Extended model  : no TP; MFE to stop or EOD; extended_r = mfe_pts / risk_pts.
    Returns outcome dict, or None if no fill.
    """
    eod_time  = dtime(EOD_HOUR, EOD_MINUTE)
    entry     = fvg['entry']
    stop      = fvg['stop']
    risk_pts  = fvg['risk_pts']
    direction = fvg['trade_dir']
    c3_dt     = fvg['c3_dt']

    tp = entry * (1.0 + TP_BPS) if direction == 'LONG' else entry * (1.0 - TP_BPS)

    fill_window = grp_df[
        (grp_df['datetime'] > c3_dt) &
        (grp_df['datetime'].dt.time <= eod_time)
    ].reset_index(drop=True)

    if fill_window.empty:
        return None

    # Locate fill bar
    fill_row = None
    for _, bar in fill_window.iterrows():
        if direction == 'LONG'  and bar['low']  <= entry:
            fill_row = bar; break
        if direction == 'SHORT' and bar['high'] >= entry:
            fill_row = bar; break

    if fill_row is None:
        return None

    fill_time = fill_row['datetime']

    # Scan from fill bar onward
    trade_bars = grp_df[
        (grp_df['datetime'] >= fill_time) &
        (grp_df['datetime'].dt.time <= eod_time)
    ].reset_index(drop=True)

    best_price  = entry   # tracks MFE
    worst_price = entry   # tracks MAE
    stop_hit    = False
    stop_hit_dt = None
    cashflow_win = False

    for _, bar in trade_bars.iterrows():
        h = bar['high']; l = bar['low']
        if direction == 'LONG':
            if l <= stop:
                worst_price = min(worst_price, stop)
                stop_hit = True; stop_hit_dt = bar['datetime']
                break
            best_price  = max(best_price, h)
            worst_price = min(worst_price, l)
            if not cashflow_win and h >= tp:
                cashflow_win = True
        else:
            if h >= stop:
                worst_price = max(worst_price, stop)
                stop_hit = True; stop_hit_dt = bar['datetime']
                break
            best_price  = min(best_price, l)
            worst_price = max(worst_price, h)
            if not cashflow_win and l <= tp:
                cashflow_win = True

    if direction == 'LONG':
        mfe_pts = max(0.0, best_price  - entry)
        mae_pts = max(0.0, entry - worst_price)
    else:
        mfe_pts = max(0.0, entry - best_price)
        mae_pts = max(0.0, worst_price - entry)

    mfe_pct = round(mfe_pts / entry * 100, 6)
    mae_pct = round(mae_pts / entry * 100, 6)

    extended_r  = round(mfe_pts / risk_pts, 4) if risk_pts > 0 else 0
    if cashflow_win:
        cashflow_r = 1.0
    elif stop_hit:
        cashflow_r = -1.0
    else:
        cashflow_r = None  # expired — neither TP nor stop reached

    return {
        'fill_time':        str(fill_time.time()),
        'cashflow_win':     'YES' if cashflow_win else 'NO',
        'cashflow_r':       cashflow_r,
        'extended_mfe_pts': round(mfe_pts, 4),
        'extended_mfe_pct': mfe_pct,
        'extended_r':       extended_r,
        'mae_pts':          round(mae_pts, 4),
        'mae_pct':          mae_pct,
        'stop_hit':         'YES' if stop_hit else 'NO',
        'stop_hit_time':    str(stop_hit_dt.time()) if stop_hit_dt else '',
    }


# ══════════════════════════════════════════════════════════════════════════════
# MAIN DAY LOOP
# ══════════════════════════════════════════════════════════════════════════════
def run_collection(df_m1, date_range=None):
    df_m1 = df_m1.copy()
    df_m1['date'] = df_m1['datetime'].dt.date

    if date_range:
        s, e = date_range
        df_m1 = df_m1[(df_m1['date'] >= s) & (df_m1['date'] <= e)]

    results = []

    for day, grp in df_m1.groupby('date'):
        grp = grp.reset_index(drop=True)

        # ── Classify day ────────────────────────────────────────────────────
        day_class, class_reason = classify_day(grp)

        row = {
            'date':               str(day),
            'day_classification': day_class,
            'class_reason':       class_reason,
            'trade':              False,
            'no_trade_reason':    '',
            'fvg_dir':            '',
            'fvg_zone_low':       None,
            'fvg_zone_high':      None,
            'fvg_c3_time':        '',
            'fvg_size_pct':       None,
            'trade_dir':          '',
            'entry_price':        None,
            'stop_price':         None,
            'risk_pts':           None,
            'risk_pct':           None,
            'fill_time':          '',
            'cashflow_win':       '',
            'cashflow_r':         None,
            'extended_mfe_pts':   None,
            'extended_mfe_pct':   None,
            'extended_r':         None,
            'mae_pts':            None,
            'mae_pct':            None,
            'stop_hit':           '',
            'stop_hit_time':      '',
        }

        # ── FVG scan — C2 must fall in [9:31, 9:59] ─────────────────────────
        # Include the 9:30 bar so it can serve as C1
        window = grp[grp['datetime'].dt.time >= dtime(9, 30)].reset_index(drop=True)

        if len(window) < 3:
            row['no_trade_reason'] = 'INSUFFICIENT BARS IN WINDOW'
            results.append(row); continue

        fvg = find_first_fvg(window.to_dict('records'))
        if fvg is None:
            row['no_trade_reason'] = 'NO FVG FOUND'
            results.append(row); continue

        row.update({
            'fvg_dir':       fvg['direction'],
            'fvg_zone_low':  fvg['zone_low'],
            'fvg_zone_high': fvg['zone_high'],
            'fvg_c3_time':   str(fvg['c3_dt'].time()),
            'fvg_size_pct':  fvg['fvg_size_pct'],
            'trade_dir':     fvg['trade_dir'],
            'entry_price':   fvg['entry'],
            'stop_price':    fvg['stop'],
            'risk_pts':      fvg['risk_pts'],
            'risk_pct':      round(fvg['risk_pts'] / fvg['entry'] * 100, 6),
        })

        outcomes = calc_fill_and_outcomes(fvg, grp)
        if outcomes is None:
            row['no_trade_reason'] = 'NO FILL IN WINDOW'
            results.append(row); continue

        row.update(outcomes)
        row['trade'] = True
        row['no_trade_reason'] = ''
        results.append(row)

    trades = [r for r in results if r['trade']]
    log.info(f"Days processed: {len(results)} | Trades: {len(trades)} | "
             f"No-trade: {len(results)-len(trades)}")
    log.info("  Classification breakdown:")
    for cls in ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']:
        cnt = sum(1 for r in results if r['day_classification'] == cls)
        pct = round(cnt / len(results) * 100, 1) if results else 0
        log.info(f"    {cls:>14}: {cnt:>5}  ({pct}%)")

    return results


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL COLUMN DEFINITIONS (shared across all sheets)
# ══════════════════════════════════════════════════════════════════════════════
HEADERS = [
    'Date', 'Classification', 'Class Reason',
    'Trade?', 'No-Trade Reason',
    'FVG Dir', 'FVG Low', 'FVG High', 'FVG C3 Time', 'FVG Size%',
    'Trade Dir', 'Entry', 'Stop', 'Risk Pts', 'Risk %',
    'Fill Time',
    'CF Win?', 'CF R',
    'Ext MFE Pts', 'Ext MFE %', 'Ext R',
    'MAE Pts', 'MAE %',
    'Stop Hit', 'Stop Time',
]
COL_KEYS = [
    'date', 'day_classification', 'class_reason',
    'trade', 'no_trade_reason',
    'fvg_dir', 'fvg_zone_low', 'fvg_zone_high', 'fvg_c3_time', 'fvg_size_pct',
    'trade_dir', 'entry_price', 'stop_price', 'risk_pts', 'risk_pct',
    'fill_time',
    'cashflow_win', 'cashflow_r',
    'extended_mfe_pts', 'extended_mfe_pct', 'extended_r',
    'mae_pts', 'mae_pct',
    'stop_hit', 'stop_hit_time',
]
COL_WIDTHS = [
    13, 16, 38,
    8,  28,
    10, 12, 12, 13, 12,
    10, 13, 13, 10, 10,
    12,
    9,  9,
    13, 12, 9,
    11, 11,
    10, 12,
]
NCOLS = len(HEADERS)


# ══════════════════════════════════════════════════════════════════════════════
# WRITE DATA ROWS (shared)
# ══════════════════════════════════════════════════════════════════════════════
def write_data_rows(ws, rows, start_row=4):
    for r_idx, row in enumerate(rows):
        er       = r_idx + start_row
        is_trade = row['trade']
        bg       = RAISED_BG if r_idx % 2 == 0 else CARD_BG
        cls_clr  = _class_color(row.get('day_classification', ''))

        for c, key in enumerate(COL_KEYS, 1):
            val  = row.get(key)
            fc   = WHITE
            bold = False
            fmt  = None

            if key == 'trade':
                val  = 'YES' if val else 'NO'
                fc   = TEAL  if val == 'YES' else RED_CLR
                bold = True
            elif key == 'day_classification':
                fc = cls_clr; bold = True
            elif key == 'class_reason':
                fc = MUTED
            elif key == 'no_trade_reason' and val:
                fc = ORANGE
            elif key in ('fvg_dir', 'trade_dir'):
                if val in ('BULLISH', 'LONG'):    fc = TEAL
                elif val in ('BEARISH', 'SHORT'): fc = RED_CLR
                bold = bool(val)
            elif key == 'cashflow_win':
                fc = TEAL if val == 'YES' else (RED_CLR if val == 'NO' else MUTED)
                bold = val == 'YES'
            elif key == 'cashflow_r':
                if val is not None:
                    fc   = TEAL if (isinstance(val, (int, float)) and val > 0) else RED_CLR
                    bold = True
                    fmt  = '0.00'
            elif key == 'stop_hit':
                fc = RED_CLR if val == 'YES' else TEAL
                bold = val == 'YES'
            elif key in ('fvg_zone_low', 'fvg_zone_high', 'entry_price', 'stop_price'):
                fmt = '#,##0.00'
            elif key in ('risk_pts', 'extended_mfe_pts', 'mae_pts'):
                fmt  = '#,##0.00'
                fc   = GOLD
                bold = is_trade
            elif key in ('fvg_size_pct', 'risk_pct', 'extended_mfe_pct', 'mae_pct'):
                fmt  = '0.0000"%"'
                fc   = GOLD
                bold = is_trade
            elif key in ('extended_r',):
                fmt  = '0.00'
                fc   = TEAL if (is_trade and val is not None and val > 0) else WHITE
                bold = is_trade

            wc(ws, er, c, val,
               font=Font(name='Arial', size=9, color=fc, bold=bold),
               fill_=fill(bg), align=center(), border=tborder(), fmt=fmt)
        ws.row_dimensions[er].height = 15


# ══════════════════════════════════════════════════════════════════════════════
# STATS BLOCK (shared)
# ══════════════════════════════════════════════════════════════════════════════
def write_stats_block(ws_stats, rows, label, clr):
    ws_stats.sheet_view.showGridLines = False
    trades    = [r for r in rows if r['trade']]
    no_trades = [r for r in rows if not r['trade']]

    ws_stats.merge_cells('A1:B1')
    wc(ws_stats, 1, 1, f'STATISTICS — {label}',
       font=hf(12, True, clr), fill_=fill(DARK_BG), align=center())
    ws_stats.row_dimensions[1].height = 28
    wcol(ws_stats, 1, 34); wcol(ws_stats, 2, 18)

    wc(ws_stats, 2, 1, 'Metric', font=hf(9, True, DARK_BG),
       fill_=fill(clr), align=center(), border=tborder(DARK_BG))
    wc(ws_stats, 2, 2, 'Value',  font=hf(9, True, DARK_BG),
       fill_=fill(clr), align=center(), border=tborder(DARK_BG))
    ws_stats.row_dimensions[2].height = 18

    def stat_row(ws, r, lbl, val, section=False, pct_fmt=False, gold=False):
        bg   = '1C2333' if section else (RAISED_BG if r % 2 == 0 else CARD_BG)
        lfc  = GOLD     if section else MUTED
        vfc  = TEAL     if gold    else (GOLD if section else WHITE)
        fmt  = '0.000000"%"' if pct_fmt else None
        wc(ws, r, 1, lbl,
           font=Font(name='Arial', size=9, bold=section, color=lfc),
           fill_=fill(bg), align=left(), border=tborder())
        wc(ws, r, 2, val,
           font=Font(name='Arial', size=9, bold=(section or gold), color=vfc),
           fill_=fill(bg), align=center(), border=tborder(), fmt=fmt)
        ws.row_dimensions[r].height = 15

    r = 3

    if not trades:
        wc(ws_stats, r, 1, 'No trades found for this classification.',
           font=cf(10, False, MUTED), fill_=fill(DARK_BG), align=left())
        ws_stats.merge_cells(f'A{r}:B{r}')
        ws_stats.freeze_panes = 'A3'
        return

    mae_vals   = [x['mae_pct']          for x in trades if x['mae_pct']          is not None]
    mfe_vals   = [x['extended_mfe_pct'] for x in trades if x['extended_mfe_pct'] is not None]
    ext_r_vals = [x['extended_r']       for x in trades if x['extended_r']       is not None]
    cf_r_vals  = [x['cashflow_r']       for x in trades if x['cashflow_r']       is not None]
    risk_vals  = [x['risk_pts']         for x in trades if x['risk_pts']         is not None]
    stop_hits  = sum(1 for x in trades if x['stop_hit']     == 'YES')
    cf_wins    = sum(1 for x in trades if x['cashflow_win'] == 'YES')
    long_cnt   = sum(1 for x in trades if x['trade_dir'] == 'LONG')
    short_cnt  = sum(1 for x in trades if x['trade_dir'] == 'SHORT')

    def pct(v, p): return round(float(np.percentile(v, p)), 6) if v else 0
    def avg(v):    return round(sum(v)/len(v), 6) if v else 0
    def med(v):    return round(float(np.median(v)), 6) if v else 0

    stat_row(ws_stats, r, '-- OVERVIEW --', '', section=True); r += 1
    stat_row(ws_stats, r, 'Total Days (this class)', len(rows));       r += 1
    stat_row(ws_stats, r, 'Total Trades',            len(trades));     r += 1
    stat_row(ws_stats, r, 'No-Trade Days',           len(no_trades));  r += 1
    stat_row(ws_stats, r, 'Trade Rate %',
             round(len(trades)/len(rows)*100, 2) if rows else 0);      r += 1
    stat_row(ws_stats, r, 'LONG Trades',             long_cnt);        r += 1
    stat_row(ws_stats, r, 'SHORT Trades',            short_cnt);       r += 1
    stat_row(ws_stats, r, 'Stop Hit Count',          stop_hits);       r += 1
    stat_row(ws_stats, r, 'Stop Hit %',
             round(stop_hits/len(trades)*100, 2) if trades else 0);    r += 1

    r += 1
    stat_row(ws_stats, r, '-- CASHFLOW MODEL  (TP = 10 bps) --', '', section=True); r += 1
    stat_row(ws_stats, r, 'CF Wins',
             cf_wins);                                                             r += 1
    stat_row(ws_stats, r, 'CF Win %',
             round(cf_wins/len(trades)*100, 2) if trades else 0);                 r += 1
    stat_row(ws_stats, r, 'Avg CF R (resolved only)',
             round(avg(cf_r_vals), 4), gold=True);                                r += 1
    stat_row(ws_stats, r, 'Sum CF R',
             round(sum(cf_r_vals), 4) if cf_r_vals else 0, gold=True);            r += 1

    r += 1
    stat_row(ws_stats, r, '-- EXTENDED MODEL  (Hold to Stop / EOD) --', '', section=True); r += 1
    stat_row(ws_stats, r, 'Avg Extended R',  round(avg(ext_r_vals), 4), gold=True); r += 1
    stat_row(ws_stats, r, 'Sum Extended R',  round(sum(ext_r_vals), 4) if ext_r_vals else 0, gold=True); r += 1
    stat_row(ws_stats, r, 'Median Ext R',    round(med(ext_r_vals), 4));             r += 1
    stat_row(ws_stats, r, 'Ext R 70th %ile', round(pct(ext_r_vals, 70), 4));        r += 1
    stat_row(ws_stats, r, 'Ext R 85th %ile', round(pct(ext_r_vals, 85), 4));        r += 1

    r += 1
    stat_row(ws_stats, r, '-- MAE %  (Risk vs Entry) --', '', section=True); r += 1
    stat_row(ws_stats, r, 'Avg MAE %',     avg(mae_vals),      pct_fmt=True, gold=True); r += 1
    stat_row(ws_stats, r, 'Median MAE %',  med(mae_vals),      pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'Max MAE %',     pct(mae_vals, 100), pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'MAE 70th %ile', pct(mae_vals, 70),  pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'MAE 90th %ile', pct(mae_vals, 90),  pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'MAE 95th %ile', pct(mae_vals, 95),  pct_fmt=True);            r += 1

    r += 1
    stat_row(ws_stats, r, '-- MFE %  (Max Favorable Move) --', '', section=True); r += 1
    stat_row(ws_stats, r, 'Avg MFE %',     avg(mfe_vals),      pct_fmt=True, gold=True); r += 1
    stat_row(ws_stats, r, 'Median MFE %',  med(mfe_vals),      pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'Max MFE %',     pct(mfe_vals, 100), pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'MFE 70th %ile', pct(mfe_vals, 70),  pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'MFE 85th %ile', pct(mfe_vals, 85),  pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'MFE 90th %ile', pct(mfe_vals, 90),  pct_fmt=True);            r += 1
    stat_row(ws_stats, r, 'MFE 95th %ile', pct(mfe_vals, 95),  pct_fmt=True);            r += 1

    r += 1
    stat_row(ws_stats, r, '-- E-RATIO  (MFE / MAE) --', '', section=True); r += 1
    e_ratio = round(avg(mfe_vals) / avg(mae_vals), 4) if avg(mae_vals) > 0 else 0
    stat_row(ws_stats, r, 'E-Ratio (Avg MFE / Avg MAE)', e_ratio, gold=True); r += 1
    stat_row(ws_stats, r, 'Sum MFE / Sum MAE',
             round(sum(mfe_vals)/sum(mae_vals), 4) if sum(mae_vals) > 0 else 0,
             gold=True); r += 1

    ws_stats.freeze_panes = 'A3'


# ══════════════════════════════════════════════════════════════════════════════
# BUILD ONE CLASSIFICATION SHEET PAIR
# ══════════════════════════════════════════════════════════════════════════════
def build_classification_sheet(wb, cls, rows):
    clr    = _class_color(cls)
    label  = CLASS_LABELS[cls]
    trades = [r for r in rows if r['trade']]

    # Data sheet
    ws = wb.create_sheet(f'{cls} Data')
    ws.sheet_view.showGridLines  = False
    ws.sheet_properties.tabColor = clr

    ws.merge_cells(f'A1:{get_column_letter(NCOLS)}1')
    wc(ws, 1, 1, label,
       font=hf(13, True, clr), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{get_column_letter(NCOLS)}2')
    subtitle = (
        f'Total: {len(rows)}  |  Trades: {len(trades)}  |  '
        f'No-Trade: {len(rows)-len(trades)}  |  '
        f'Trade Rate: {len(trades)/len(rows)*100:.1f}%'
        if rows else 'No data'
    )
    wc(ws, 2, 1, subtitle,
       font=cf(10, False, MUTED), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[2].height = 18

    for c, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        wc(ws, 3, c, h,
           font=hf(9, True, DARK_BG), fill_=fill(clr),
           align=center(), border=tborder(DARK_BG))
        wcol(ws, c, w)
    ws.row_dimensions[3].height = 20

    write_data_rows(ws, rows, start_row=4)
    ws.freeze_panes = 'A4'

    # Stats sheet
    ws_stats = wb.create_sheet(f'{cls} Stats')
    ws_stats.sheet_properties.tabColor = clr
    write_stats_block(ws_stats, rows, label, clr)


# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_summary_sheet(wb, results):
    ws = wb.active
    ws.title = 'Summary'
    ws.sheet_view.showGridLines  = False
    ws.sheet_properties.tabColor = GOLD

    total_days   = len(results)
    total_trades = sum(1 for r in results if r['trade'])

    ws.merge_cells('A1:I1')
    wc(ws, 1, 1, 'FVG PHASE 1 v2 — Daily Classification Overview',
       font=hf(14, True, GOLD), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:I2')
    wc(ws, 2, 1,
       (f'Total Days: {total_days}  |  Total Trades: {total_trades}  |  '
        f'Overall Trade Rate: {total_trades/total_days*100:.1f}%'
        if total_days else 'No data'),
       font=cf(10, False, MUTED), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[2].height = 18

    hdrs   = ['Classification', 'Days', 'Trades', 'No-Trade',
              'Trade Rate %', 'Avg MAE %', 'Avg MFE %', 'E-Ratio', 'CF Win %']
    widths = [22, 10, 10, 12, 14, 14, 14, 12, 12]

    for c, (h, w) in enumerate(zip(hdrs, widths), 1):
        wc(ws, 3, c, h,
           font=hf(9, True, DARK_BG), fill_=fill(GOLD),
           align=center(), border=tborder(DARK_BG))
        wcol(ws, c, w)
    ws.row_dimensions[3].height = 20

    for ri, cls in enumerate(['DWP', 'DNP', 'R1', 'R2', 'Unclassified']):
        er  = ri + 4
        bg  = RAISED_BG if ri % 2 == 0 else CARD_BG
        clr = _class_color(cls)

        cls_rows   = [r for r in results if r['day_classification'] == cls]
        cls_trades = [r for r in cls_rows  if r['trade']]

        mae_vals = [r['mae_pct']          for r in cls_trades if r['mae_pct']          is not None]
        mfe_vals = [r['extended_mfe_pct'] for r in cls_trades if r['extended_mfe_pct'] is not None]
        cf_wins  = sum(1 for r in cls_trades if r['cashflow_win'] == 'YES')

        avg_mae  = round(sum(mae_vals)/len(mae_vals), 4) if mae_vals else 0
        avg_mfe  = round(sum(mfe_vals)/len(mfe_vals), 4) if mfe_vals else 0
        e_ratio  = round(avg_mfe / avg_mae, 4) if avg_mae > 0 else 0
        tr_rate  = round(len(cls_trades)/len(cls_rows)*100, 1) if cls_rows else 0
        cf_pct   = round(cf_wins/len(cls_trades)*100, 1) if cls_trades else 0

        vals = [cls, len(cls_rows), len(cls_trades),
                len(cls_rows)-len(cls_trades), tr_rate,
                avg_mae, avg_mfe, e_ratio, cf_pct]
        fmts = [None, None, None, None,
                '0.0"%"', '0.0000"%"', '0.0000"%"', '0.0000', '0.0"%"']

        for c, (val, fmt) in enumerate(zip(vals, fmts), 1):
            fc   = clr if c == 1 else (GOLD if c in (6, 7, 8) else WHITE)
            bold = (c == 1)
            wc(ws, er, c, val,
               font=Font(name='Arial', size=10, color=fc, bold=bold),
               fill_=fill(bg), align=center(), border=tborder(), fmt=fmt)
        ws.row_dimensions[er].height = 18

    ws.freeze_panes = 'A4'


# ══════════════════════════════════════════════════════════════════════════════
# NO-TRADE BREAKDOWN SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_notrade_sheet(wb, results):
    ws2 = wb.create_sheet('No-Trade Breakdown')
    ws2.sheet_view.showGridLines  = False
    ws2.sheet_properties.tabColor = RED_CLR

    ws2.merge_cells('A1:C1')
    wc(ws2, 1, 1, 'NO-TRADE BREAKDOWN — Dropout Funnel',
       font=hf(12, True, RED_CLR), fill_=fill(DARK_BG), align=center())
    ws2.row_dimensions[1].height = 28

    no_trades = [r for r in results if not r['trade']]
    reasons   = {}
    for r in no_trades:
        reason = r['no_trade_reason'] or 'UNKNOWN'
        reasons[reason] = reasons.get(reason, 0) + 1

    funnel_order = [
        'INSUFFICIENT BARS IN WINDOW',
        'NO FVG FOUND',
        'NO FILL IN WINDOW',
    ]

    wc(ws2, 2, 1, 'Reason',          font=hf(9, True, DARK_BG), fill_=fill(RED_CLR), align=center(), border=tborder(DARK_BG))
    wc(ws2, 2, 2, 'Count',           font=hf(9, True, DARK_BG), fill_=fill(RED_CLR), align=center(), border=tborder(DARK_BG))
    wc(ws2, 2, 3, '% of Total Days', font=hf(9, True, DARK_BG), fill_=fill(RED_CLR), align=center(), border=tborder(DARK_BG))
    wcol(ws2, 1, 38); wcol(ws2, 2, 10); wcol(ws2, 3, 18)
    ws2.row_dimensions[2].height = 20

    total_days = len(results)
    ri = 3
    for reason in funnel_order:
        cnt = reasons.get(reason, 0)
        p   = round(cnt/total_days*100, 1) if total_days else 0
        bg  = RAISED_BG if ri % 2 == 0 else CARD_BG
        wc(ws2, ri, 1, reason, font=cf(9, False, ORANGE), fill_=fill(bg), align=left(),   border=tborder())
        wc(ws2, ri, 2, cnt,    font=hf(9, True,  WHITE),  fill_=fill(bg), align=center(), border=tborder())
        wc(ws2, ri, 3, p,      font=cf(9, False, MUTED),  fill_=fill(bg), align=center(), border=tborder(), fmt='0.0"%"')
        ws2.row_dimensions[ri].height = 15
        ri += 1

    for reason, cnt in reasons.items():
        if reason not in funnel_order:
            p  = round(cnt/total_days*100, 1) if total_days else 0
            bg = RAISED_BG if ri % 2 == 0 else CARD_BG
            wc(ws2, ri, 1, reason, font=cf(9, False, ORANGE), fill_=fill(bg), align=left(),   border=tborder())
            wc(ws2, ri, 2, cnt,    font=hf(9, True,  WHITE),  fill_=fill(bg), align=center(), border=tborder())
            wc(ws2, ri, 3, p,      font=cf(9, False, MUTED),  fill_=fill(bg), align=center(), border=tborder(), fmt='0.0"%"')
            ws2.row_dimensions[ri].height = 15
            ri += 1

    ws2.row_dimensions[ri].height = 4; ri += 1
    trade_cnt = sum(1 for r in results if r['trade'])
    wc(ws2, ri, 1, 'TRADE RATE',
       font=hf(9, True, TEAL), fill_=fill(RAISED_BG), align=left(), border=tborder())
    wc(ws2, ri, 2, trade_cnt,
       font=hf(9, True, TEAL), fill_=fill(RAISED_BG), align=center(), border=tborder())
    wc(ws2, ri, 3, round(trade_cnt/total_days*100, 1) if total_days else 0,
       font=hf(9, True, TEAL), fill_=fill(RAISED_BG), align=center(),
       border=tborder(), fmt='0.0"%"')
    ws2.row_dimensions[ri].height = 16
    ws2.freeze_panes = 'A3'


# ══════════════════════════════════════════════════════════════════════════════
# UNCLASSIFIED DEBUG SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_unclassified_debug_sheet(wb, results):
    """
    Two-part sheet:
      Top    — reason summary table (count + % per unique reason)
      Bottom — full per-day detail (date, reason, RTH bar count, hours present)
    """
    unclassified = [r for r in results if r['day_classification'] == 'Unclassified']
    total_days   = len(results)

    ws = wb.create_sheet('Unclassified Debug')
    ws.sheet_view.showGridLines  = False
    ws.sheet_properties.tabColor = MUTED

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells('A1:E1')
    wc(ws, 1, 1, f'UNCLASSIFIED DEBUG — {len(unclassified)} days  ({round(len(unclassified)/total_days*100,1)}% of total)',
       font=hf(12, True, ORANGE), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 28

    # ── Section 1: Reason summary ────────────────────────────────────────────
    ws.merge_cells('A2:E2')
    wc(ws, 2, 1, 'REASON SUMMARY',
       font=hf(10, True, GOLD), fill_=fill(RAISED_BG), align=center())
    ws.row_dimensions[2].height = 20

    sum_headers = ['Reason', 'Count', '% of Unclassified', '% of All Days']
    sum_widths  = [52, 10, 20, 16]
    for c, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        wc(ws, 3, c, h,
           font=hf(9, True, DARK_BG), fill_=fill(ORANGE),
           align=center(), border=tborder(DARK_BG))
        wcol(ws, c, w)
    ws.row_dimensions[3].height = 18

    # Count by reason, sorted descending
    from collections import Counter
    reason_counts = Counter(r['class_reason'] for r in unclassified)
    sorted_reasons = reason_counts.most_common()

    ri = 4
    for reason, cnt in sorted_reasons:
        bg      = RAISED_BG if ri % 2 == 0 else CARD_BG
        pct_u   = round(cnt / len(unclassified) * 100, 1) if unclassified else 0
        pct_all = round(cnt / total_days        * 100, 1) if total_days   else 0
        wc(ws, ri, 1, reason, font=cf(9, False, WHITE),  fill_=fill(bg), align=left(),   border=tborder())
        wc(ws, ri, 2, cnt,    font=hf(9, True,  GOLD),   fill_=fill(bg), align=center(), border=tborder())
        wc(ws, ri, 3, pct_u,  font=cf(9, False, MUTED),  fill_=fill(bg), align=center(), border=tborder(), fmt='0.0"%"')
        wc(ws, ri, 4, pct_all,font=cf(9, False, MUTED),  fill_=fill(bg), align=center(), border=tborder(), fmt='0.0"%"')
        ws.row_dimensions[ri].height = 15
        ri += 1

    # ── Spacer ───────────────────────────────────────────────────────────────
    ri += 1
    ws.row_dimensions[ri].height = 8
    ri += 1

    # ── Section 2: Per-day detail ────────────────────────────────────────────
    ws.merge_cells(f'A{ri}:E{ri}')
    wc(ws, ri, 1, 'PER-DAY DETAIL',
       font=hf(10, True, GOLD), fill_=fill(RAISED_BG), align=center())
    ws.row_dimensions[ri].height = 20
    ri += 1

    detail_headers = ['Date', 'Classification Reason', 'FVG No-Trade Reason',
                      'Trade?', 'FVG Dir']
    detail_widths  = [14, 52, 32, 10, 12]
    for c, (h, w) in enumerate(zip(detail_headers, detail_widths), 1):
        wc(ws, ri, c, h,
           font=hf(9, True, DARK_BG), fill_=fill(ORANGE),
           align=center(), border=tborder(DARK_BG))
        wcol(ws, c, w)
    ws.row_dimensions[ri].height = 18
    ri += 1

    for idx, row in enumerate(unclassified):
        bg         = RAISED_BG if idx % 2 == 0 else CARD_BG
        trade_val  = 'YES' if row['trade'] else 'NO'
        trade_fc   = TEAL if row['trade'] else RED_CLR
        fvg1_fc    = TEAL if row.get('fvg_dir') == 'BULLISH' else (
                     RED_CLR if row.get('fvg_dir') == 'BEARISH' else MUTED)

        wc(ws, ri, 1, row['date'],
           font=cf(9, False, WHITE),  fill_=fill(bg), align=center(), border=tborder())
        wc(ws, ri, 2, row['class_reason'],
           font=cf(9, False, ORANGE), fill_=fill(bg), align=left(),   border=tborder())
        wc(ws, ri, 3, row.get('no_trade_reason') or '—',
           font=cf(9, False, MUTED),  fill_=fill(bg), align=left(),   border=tborder())
        wc(ws, ri, 4, trade_val,
           font=Font(name='Arial', size=9, bold=True, color=trade_fc),
           fill_=fill(bg), align=center(), border=tborder())
        wc(ws, ri, 5, row.get('fvg_dir') or '—',
           font=Font(name='Arial', size=9, bold=bool(row.get('fvg_dir')), color=fvg1_fc),
           fill_=fill(bg), align=center(), border=tborder())
        ws.row_dimensions[ri].height = 15
        ri += 1

    ws.freeze_panes = 'A4'


# ══════════════════════════════════════════════════════════════════════════════
# MASTER BUILD
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(results, out_dir):
    """
    Sheet order:
      1.  Summary                 (gold tab)
      2.  DWP Data                (teal tab)
      3.  DWP Stats
      4.  DNP Data                (gold tab)
      5.  DNP Stats
      6.  R1 Data                 (blue tab)
      7.  R1 Stats
      8.  R2 Data                 (orange tab)
      9.  R2 Stats
      10. Unclassified Data       (grey tab)
      11. Unclassified Stats
      12. No-Trade Breakdown      (red tab)
      13. Unclassified Debug      (grey tab) — reason summary + per-day detail
    """
    wb = Workbook()

    build_summary_sheet(wb, results)

    for cls in ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']:
        cls_rows = [r for r in results if r['day_classification'] == cls]
        build_classification_sheet(wb, cls, cls_rows)

    build_notrade_sheet(wb, results)
    build_unclassified_debug_sheet(wb, results)

    ts    = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f'FVG_Phase1_v2_{ts}.xlsx'
    path  = os.path.join(out_dir, fname)
    wb.save(path)
    log.info(f"Saved: {path}")
    return path


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print()
    print('=' * 60)
    print('  FVG PHASE 1 v2 — Classification + MAE/MFE Collector')
    print('=' * 60)
    print()
    print('  Day types    : DWP | DNP | R1 | R2 | Unclassified')
    print('  FVG window   : 09:31 to 09:59 ET (C2 in window)')
    print('  Entry        : Long @ C3.low | Short @ C3.high')
    print('  Stop         : min/max(C1, C2) extremes')
    print('  Fill window  : C3+1 bar through 16:00 ET')
    print('  Cashflow TP  : 10 bps (0.10%)')
    print('  Extended     : hold to stop or EOD')
    print()

    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--table', default='nq_1m', help='DuckDB table (nq_1m or es_1m)')
    args, _ = parser.parse_known_args()
    log.info(f'Loading {args.table} from DuckDB...')
    df_m1 = load_ohlc(args.table)

    # Date range
    print('\nDate range filter (leave blank for all data):')
    date_range = None
    try:
        s = input('  Start (YYYY-MM-DD or blank): ').strip()
        e = input('  End   (YYYY-MM-DD or blank): ').strip()
        if s and e:
            from datetime import date
            date_range = (datetime.strptime(s, '%Y-%m-%d').date(),
                          datetime.strptime(e, '%Y-%m-%d').date())
            log.info(f'  Date range: {date_range[0]} -> {date_range[1]}')
        else:
            log.info('  No date range — full dataset')
    except Exception:
        log.info('  Invalid date — full dataset')

    # Run
    print()
    log.info('Running classification + setup detection...')
    results = run_collection(df_m1, date_range=date_range)

    out_dir = os.path.dirname(os.path.abspath(__file__))
    path    = build_excel(results, out_dir)

    trades = [r for r in results if r['trade']]
    print()
    print('=' * 60)
    print('  DONE')
    print(f'  Total days   : {len(results)}')
    print(f'  Trades       : {len(trades)}')
    print(f'  No-trade     : {len(results)-len(trades)}')
    print()
    print('  Classification breakdown:')
    for cls in ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']:
        cnt = sum(1 for r in results if r['day_classification'] == cls)
        pct = round(cnt/len(results)*100, 1) if results else 0
        print(f'    {cls:>14} : {cnt:>5}  ({pct}%)')
    if trades:
        mae_vals = [r['mae_pct']          for r in trades if r['mae_pct']          is not None]
        mfe_vals = [r['extended_mfe_pct'] for r in trades if r['extended_mfe_pct'] is not None]
        cf_wins  = sum(1 for r in trades if r['cashflow_win'] == 'YES')
        if mae_vals:
            print()
            print(f'  Avg MAE%     : {sum(mae_vals)/len(mae_vals):.4f}%')
            print(f'  Avg MFE%     : {sum(mfe_vals)/len(mfe_vals):.4f}%')
            print(f'  E-Ratio      : {sum(mfe_vals)/sum(mae_vals):.4f}')
            print(f'  CF Win %     : {cf_wins/len(trades)*100:.1f}%')
    print(f'  Output       : {path}')
    print('=' * 60)


if __name__ == '__main__':
    main()
