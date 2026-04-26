"""
Build the forward-test tracker as a ready-to-use .xlsx file.

Run from this folder:
    python3 build_tracker.py

Output: forward_test_tracker.xlsx with two pre-formatted sheets (Trades + Summary)
and an embedded equity-curve chart. All formulas pre-populated for the first
500 trade rows; just import to Google Sheets or open in Excel and start logging.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).parent / "forward_test_tracker.xlsx"

# How many empty rows to pre-populate with row formulas.
# 500 covers ~3.5 years of trading at the 1H_5M model's frequency.
N_ROWS = 500

# Column layout for the Trades sheet
TRADES_COLS = [
    ("trade_no",         12, "Sequence number (1, 2, 3 …)"),
    ("date",             12, "Trade date (YYYY-MM-DD)"),
    ("time_et",          10, "Entry time in ET (HH:MM)"),
    ("combo",            10, "1H/5M or 30M/3M"),
    ("direction",        10, "LONG or SHORT"),
    ("smt",              8,  "TRUE / FALSE — was SMT divergence true at entry"),
    ("planned_entry",    14, "Entry price from indicator alert"),
    ("planned_sl",       14, "Stop-loss price (sweep extreme)"),
    ("planned_tp",       14, "Take-profit price (1R from entry)"),
    ("planned_risk_pts", 16, "Auto: |planned_entry - planned_sl|"),
    ("actual_entry",     14, "Price you actually filled at"),
    ("actual_exit",      14, "Price the trade exited at (TP, SL, BE, or manual)"),
    ("outcome",          14, "WIN / LOSS / BE / SKIPPED / MANUAL_EXIT"),
    ("r_realized",       12, "Actual R captured (e.g. +1.0, -1.0, 0)"),
    ("mae_R",            10, "Deepest adverse excursion in R-units (eyeball OK)"),
    ("mfe_R",            10, "Furthest favorable excursion in R-units"),
    ("slippage_pts",     14, "actual_entry - planned_entry (signed)"),
    ("contracts",        10, "Number of MNQ contracts traded"),
    ("pnl_usd",          12, "Auto: r_realized × contracts × risk_pts × $2"),
    ("notes",            40, "Free text — context, news, fill quality, regime"),
]

# Color palette — muted, prints well, distinct enough for at-a-glance reading
HDR_FILL  = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HDR_FONT  = Font(color="F9FAFB", bold=True, size=11)
ZEBRA     = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
AUTO_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amber for auto-computed cells
GOOD_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # green for healthy stats
WARN_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
LABEL_FONT = Font(bold=True, size=11, color="111827")
HEADLINE   = Font(bold=True, size=14, color="111827")
SUBHEAD    = Font(bold=True, size=11, color="374151")
SUBTLE     = Font(size=10, color="6B7280", italic=True)

THIN = Side(border_style="thin", color="D1D5DB")
BOX  = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def col_letter(idx_1_based: int) -> str:
    return get_column_letter(idx_1_based)


def build_trades_sheet(wb):
    ws = wb.active
    ws.title = "Trades"

    # ── Header row ──────────────────────────────────────────────────────────
    for i, (name, width, _) in enumerate(TRADES_COLS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[col_letter(i)].width = width

    # Tooltip / comment on each header cell so users know what goes where
    from openpyxl.comments import Comment
    for i, (name, _, doc) in enumerate(TRADES_COLS, start=1):
        ws.cell(row=1, column=i).comment = Comment(doc, "FractalSweep")

    ws.row_dimensions[1].height = 22

    # ── Pre-populated rows ──────────────────────────────────────────────────
    # Column indices (1-based) for formula references
    COL = {name: idx for idx, (name, _, _) in enumerate(TRADES_COLS, start=1)}

    for r in range(2, 2 + N_ROWS):
        # trade_no = sequential
        ws.cell(row=r, column=COL["trade_no"], value=r - 1)

        # planned_risk_pts: auto from planned_entry - planned_sl (absolute value)
        # Empty if either is blank → the IFERROR keeps the cell clean.
        e_ref  = f"{col_letter(COL['planned_entry'])}{r}"
        sl_ref = f"{col_letter(COL['planned_sl'])}{r}"
        ws.cell(
            row=r, column=COL["planned_risk_pts"],
            value=f'=IFERROR(IF(OR({e_ref}="",{sl_ref}=""),"",ABS({e_ref}-{sl_ref})),"")',
        ).fill = AUTO_FILL

        # pnl_usd: auto from r_realized * contracts * risk_pts * $2/pt (MNQ)
        r_ref = f"{col_letter(COL['r_realized'])}{r}"
        c_ref = f"{col_letter(COL['contracts'])}{r}"
        rp_ref = f"{col_letter(COL['planned_risk_pts'])}{r}"
        ws.cell(
            row=r, column=COL["pnl_usd"],
            value=f'=IFERROR(IF(OR({r_ref}="",{c_ref}="",{rp_ref}=""),"",{r_ref}*{c_ref}*{rp_ref}*2),"")',
        ).fill = AUTO_FILL

    # Freeze the header row + the first column
    ws.freeze_panes = "B2"

    # Number formats on the columns where it helps readability
    price_fmt = "#,##0.00"
    r_fmt     = "+0.00;-0.00;0.00"
    usd_fmt   = '"$"#,##0.00'
    pct_fmt   = "0.0%"

    for r in range(2, 2 + N_ROWS):
        for col_name in ("planned_entry", "planned_sl", "planned_tp",
                          "actual_entry", "actual_exit"):
            ws.cell(row=r, column=COL[col_name]).number_format = price_fmt
        ws.cell(row=r, column=COL["planned_risk_pts"]).number_format = price_fmt
        ws.cell(row=r, column=COL["r_realized"]).number_format       = r_fmt
        ws.cell(row=r, column=COL["mae_R"]).number_format            = r_fmt
        ws.cell(row=r, column=COL["mfe_R"]).number_format            = r_fmt
        ws.cell(row=r, column=COL["slippage_pts"]).number_format     = price_fmt
        ws.cell(row=r, column=COL["pnl_usd"]).number_format          = usd_fmt

    return COL


def build_summary_sheet(wb, trades_cols):
    ws = wb.create_sheet("Summary")

    # Wider columns for the labels and chart
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 4   # spacer
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    R_COL_LETTER = col_letter(trades_cols["r_realized"])
    OUTCOME_COL = col_letter(trades_cols["outcome"])
    PNL_COL     = col_letter(trades_cols["pnl_usd"])

    # ── Title ──────────────────────────────────────────────────────────────
    ws["A1"] = "Fractal Sweep — Forward-Test Summary"
    ws["A1"].font = HEADLINE
    ws.merge_cells("A1:F1")

    ws["A2"] = "Auto-updates from the Trades tab. Only WIN/LOSS rows count toward stats; SKIPPED and BE excluded from WR/EV."
    ws["A2"].font = SUBTLE
    ws.merge_cells("A2:F2")

    # ── Headline stats (rows 4-15) ─────────────────────────────────────────
    ws["A4"] = "Headline stats"
    ws["A4"].font = SUBHEAD

    headline = [
        ("Total trades logged",
         f'=COUNTA(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1})',
         "0"),
        ("Wins",
         f'=COUNTIF(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1},"WIN")',
         "0"),
        ("Losses",
         f'=COUNTIF(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1},"LOSS")',
         "0"),
        ("Breakevens",
         f'=COUNTIF(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1},"BE")',
         "0"),
        ("Skipped",
         f'=COUNTIF(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1},"SKIPPED")',
         "0"),
        ("Win rate (excl BE)",
         '=IFERROR(B6/(B6+B7),0)',
         pct_fmt := "0.0%"),
        ("Avg R per trade (EV)",
         f'=IFERROR(SUM(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1})/(B6+B7+B8),0)',
         "+0.000;-0.000;0.000"),
        ("Profit factor",
         f'=IFERROR(SUMIF(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1},">0")/ABS(SUMIF(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1},"<0")),0)',
         "0.00"),
        ("Total R",
         f'=SUM(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1})',
         "+0.00;-0.00;0.00"),
        ("Total P&L ($)",
         f'=SUM(Trades!{PNL_COL}2:{PNL_COL}{N_ROWS+1})',
         '"$"#,##0.00'),
        ("Max drawdown (R)",
         '=IFERROR(MIN(F4:F503),0)',  # F = drawdown column built below
         "0.00"),
    ]
    for i, (label, formula, fmt) in enumerate(headline, start=5):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = LABEL_FONT
        ws[f"B{i}"] = formula
        ws[f"B{i}"].number_format = fmt
        ws[f"B{i}"].alignment = Alignment(horizontal="right")

    # ── Status indicator (row 16-17) ───────────────────────────────────────
    ws["A16"] = "Forward-test status"
    ws["A16"].font = SUBHEAD

    ws["A17"] = "(auto)"
    ws["A17"].font = SUBTLE
    # Logic: <30 trades = noise; <100 = early; <45% WR or <0 EV = STOP; else on-track
    ws["B17"] = (
        '=IF(B5<30,"⏳ Noise — keep tracking",'
        'IF(B5<100,"📊 Trend — too early to call",'
        'IF(B10<0.45,"🛑 STOP — WR below 45%",'
        'IF(B11<0,"🛑 STOP — EV negative",'
        '"✓ On track"))))'
    )
    ws["B17"].font = Font(bold=True, size=12)
    ws["B17"].alignment = Alignment(horizontal="center")
    ws["B17"].fill = GOOD_FILL  # baseline; user sees it change with content

    # ── Rolling stats (rows 4-15 in column D) ──────────────────────────────
    ws["D4"] = "Rolling windows"
    ws["D4"].font = SUBHEAD

    # Strategy: build rolling WR/EV using OFFSET/COUNTA to grab the last N
    # WIN-or-LOSS outcomes. This works regardless of how many rows are filled.
    #
    # Approach: count W/L in the last N populated rows of the outcome column.
    # We derive `last_filled_row` then construct a slice.
    #
    # This is clearer with helper cells. We use D5:D6 for the sentinel.

    # Helper — row of the most recent populated trade row (for slicing)
    last_row_formula = (
        f'=IFERROR(MATCH(2,1/(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1}<>""))+1,1)'
    )
    # Entered as array formula via _ARRAY hack — this pattern works in Sheets and modern Excel
    # but to keep simple we use a different approach: use COUNTA which gives count of populated cells
    # in the outcome column — assuming outcomes are filled top-down.
    populated = (
        f'COUNTA(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1})'
    )

    rolling = [
        ("Last 30 — WR",
         f'=IFERROR(COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-30),0,MIN(30,{populated}),1),"WIN")/(COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-30),0,MIN(30,{populated}),1),"WIN")+COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-30),0,MIN(30,{populated}),1),"LOSS")),0)',
         "0.0%"),
        ("Last 30 — EV",
         f'=IFERROR(SUMPRODUCT(OFFSET(Trades!{R_COL_LETTER}2,MAX(0,{populated}-30),0,MIN(30,{populated}),1))/MIN(30,{populated}),0)',
         "+0.000;-0.000;0.000"),
        ("Last 50 — WR",
         f'=IFERROR(COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-50),0,MIN(50,{populated}),1),"WIN")/(COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-50),0,MIN(50,{populated}),1),"WIN")+COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-50),0,MIN(50,{populated}),1),"LOSS")),0)',
         "0.0%"),
        ("Last 50 — EV",
         f'=IFERROR(SUMPRODUCT(OFFSET(Trades!{R_COL_LETTER}2,MAX(0,{populated}-50),0,MIN(50,{populated}),1))/MIN(50,{populated}),0)',
         "+0.000;-0.000;0.000"),
        ("Last 100 — WR",
         f'=IFERROR(COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-100),0,MIN(100,{populated}),1),"WIN")/(COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-100),0,MIN(100,{populated}),1),"WIN")+COUNTIFS(OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-100),0,MIN(100,{populated}),1),"LOSS")),0)',
         "0.0%"),
        ("Last 100 — EV",
         f'=IFERROR(SUMPRODUCT(OFFSET(Trades!{R_COL_LETTER}2,MAX(0,{populated}-100),0,MIN(100,{populated}),1))/MIN(100,{populated}),0)',
         "+0.000;-0.000;0.000"),
    ]
    for i, (label, formula, fmt) in enumerate(rolling, start=5):
        ws[f"D{i}"] = label
        ws[f"D{i}"].font = LABEL_FONT
        ws[f"E{i}"] = formula
        ws[f"E{i}"].number_format = fmt
        ws[f"E{i}"].alignment = Alignment(horizontal="right")

    # ── Equity curve & drawdown (rows 4-503 in columns E and F) ───────────
    # We use rows past the rolling-stats area (E11+ free).
    ws["E12"] = "running_R"
    ws["E12"].font = LABEL_FONT
    ws["F12"] = "drawdown_R"
    ws["F12"].font = LABEL_FONT

    # E13 = first trade's r if it's a WIN/LOSS; F13 = 0
    ws["E13"] = f'=IF(OR(Trades!{OUTCOME_COL}2="WIN",Trades!{OUTCOME_COL}2="LOSS"),Trades!{R_COL_LETTER}2,0)'
    ws["F13"] = "=MIN(0,E13-MAX($E$13:E13))"

    # E14:E512 — running sum, only counting WIN/LOSS rows
    for offset in range(1, N_ROWS):
        excel_row = 13 + offset      # 14, 15, ..., 512
        trades_row = 2 + offset      # 3, 4, ..., 501
        prev_running = f"E{excel_row - 1}"
        ws[f"E{excel_row}"] = (
            f'=IF(OR(Trades!{OUTCOME_COL}{trades_row}="WIN",'
            f'Trades!{OUTCOME_COL}{trades_row}="LOSS"),'
            f'{prev_running}+Trades!{R_COL_LETTER}{trades_row},'
            f'{prev_running})'
        )
        ws[f"F{excel_row}"] = f"=MIN(0,E{excel_row}-MAX($E$13:E{excel_row}))"
        ws[f"E{excel_row}"].number_format = "+0.00;-0.00;0.00"
        ws[f"F{excel_row}"].number_format = "0.00"

    ws["E13"].number_format = "+0.00;-0.00;0.00"
    ws["F13"].number_format = "0.00"

    # Fix the headline Max DD reference now that we know the column
    ws["B15"] = f'=IFERROR(MIN(F13:F{12 + N_ROWS}),0)'

    # ── Equity-curve chart ─────────────────────────────────────────────────
    chart = LineChart()
    chart.title = "Equity curve (cumulative R)"
    chart.style = 12
    chart.y_axis.title = "R"
    chart.x_axis.title = "Trade #"
    chart.height = 10
    chart.width = 18
    chart.legend = None

    data = Reference(ws, min_col=5, min_row=12, max_col=5, max_row=12 + N_ROWS)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, "H4")

    # ── Footer / quick legend ──────────────────────────────────────────────
    ws["A20"] = "Decision rules"
    ws["A20"].font = SUBHEAD
    rules = [
        "🛑 Stop trading if cumulative drawdown > 15R = $3,375",
        "🛑 Stop if WR < 45% on Last-50 window",
        "🛑 Stop if EV < 0R on Last-100 window",
        "✓  Continue if WR 50-65% on rolling windows (within expected variance)",
        "✓  Continue through 5-trade losing streaks (math says they're common at 59% WR)",
    ]
    for i, rule in enumerate(rules, start=21):
        ws[f"A{i}"] = rule
        ws[f"A{i}"].font = Font(size=10, color="374151")
        ws.merge_cells(f"A{i}:F{i}")

    ws["A27"] = "Validation criteria — what 'edge confirmed' looks like after ~200 trades:"
    ws["A27"].font = SUBHEAD
    criteria = [
        "WR consistently 55–62% on rolling windows",
        "Avg R per trade between +0.10 and +0.18",
        "Max drawdown stayed within 15R",
        "Profit factor 1.4 or higher",
    ]
    for i, c in enumerate(criteria, start=28):
        ws[f"A{i}"] = "• " + c
        ws[f"A{i}"].font = Font(size=10, color="374151")
        ws.merge_cells(f"A{i}:F{i}")

    ws.freeze_panes = "A4"


def main():
    wb = Workbook()
    trades_cols = build_trades_sheet(wb)
    build_summary_sheet(wb, trades_cols)
    wb.save(OUT)
    print(f"✓ Wrote {OUT}")
    print(f"  {N_ROWS} trade rows pre-populated")
    print(f"  Open in Excel or upload to Google Sheets (File → Import) to start")


if __name__ == "__main__":
    main()
